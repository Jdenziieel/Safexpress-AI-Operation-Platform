from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import Dict, List, Any, Optional
import pandas as pd
import io
import json
from datetime import datetime
import os
import requests
import time
from functools import wraps
from dotenv import load_dotenv
import numpy as np
from safexpressops_target_columns import SAFEXPRESSOPS_TARGET_COLUMNS
from smart_mapping_engine import SmartMappingEngine
from io import StringIO

MONITORING_URL = os.getenv("MONITORING_SERVICE_URL", "")
load_dotenv()

# Import the smart mapping engine
try:
    from smart_mapping_engine import SmartMappingEngine
    SMART_MAPPING_AVAILABLE = True
except ImportError:
    print("Warning: SmartMappingEngine not found. Smart mapping will use fallback logic.")
    SMART_MAPPING_AVAILABLE = False
    
app = FastAPI(title="Mapping Agent API", version="1.0.0")

# Pydantic Models
class ToolRequest(BaseModel):
    """Generic tool execution request"""
    tool: str
    inputs: Dict[str, Any]
    credentials: Optional[Dict[str, Any]] = None


class ToolResponse(BaseModel):
    """Generic tool execution response"""
    success: bool
    result: Optional[Dict[str, Any]] = None
    error: Optional[str] = None


# In-memory storage for mapping templates (use Redis/DB in production)
MAPPING_TEMPLATES = {}


# ============================================================
# MONITORING UTILITIES (must be defined before use)
# ============================================================

def calculate_accuracy(result: Any, task_type: str) -> float:
    """Calculate task-specific accuracy score"""
    if not isinstance(result, dict):
        return 100.0

    if not result.get("success"):
        return 0.0

    # Task-specific accuracy calculation
    if task_type == "smart_column_mapping":
        high_conf = result.get("high_confidence_count", 0)
        total = len(result.get("mappings", {}))
        accuracy = (high_conf / total * 100) if total > 0 else 0
        return min(accuracy, 100.0)

    elif task_type == "parse_file":
        return 100.0 if result.get("columns") else 0

    elif task_type == "transform_data":
        return 100.0 if result.get("transformed_data") else 0

    else:
        return 100.0


def monitor_task(agent_name: str, task_type: str):
    """Decorator to monitor agent tasks"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            task_id = f"{agent_name}_{int(time.time() * 1000)}"
            start_time = time.time()

            try:
                result = func(*args, **kwargs)
                latency = time.time() - start_time
                success = result.get("success", False) if isinstance(result, dict) else True
                accuracy = calculate_accuracy(result, task_type)

                try:
                    requests.post(
                        f"{MONITORING_URL}/metrics/record",
                        json={
                            "agent_name": agent_name,
                            "task_id": task_id,
                            "timestamp": datetime.now().isoformat(),
                            "accuracy_score": accuracy,
                            "latency_seconds": latency,
                            "success": success,
                            "error_message": result.get("error") if isinstance(result, dict) else None,
                            "task_type": task_type,
                            "input_size": len(str(kwargs)) if kwargs else 0,
                            "output_size": len(str(result)) if result else 0,
                        },
                        timeout=2,
                    )
                    print(f"   Monitoring: {task_type} | Success: {success} | Accuracy: {accuracy:.1f}%")
                except Exception as e:
                    print(f"   Monitoring report failed: {str(e)}")

                return result

            except Exception as e:
                latency = time.time() - start_time
                try:
                    requests.post(
                        f"{MONITORING_URL}/metrics/record",
                        json={
                            "agent_name": agent_name,
                            "task_id": task_id,
                            "timestamp": datetime.now().isoformat(),
                            "accuracy_score": 0,
                            "latency_seconds": latency,
                            "success": False,
                            "error_message": str(e),
                            "task_type": task_type,
                        },
                        timeout=2,
                    )
                except:
                    pass
                raise

        return wrapper
    return decorator


# ============================================================
# TOOL IMPLEMENTATIONS
# ============================================================

def normalize_column_name(col_name: str) -> str:
    """
    Normalize column names by removing newlines and extra spaces
    
    Examples:
        "Discrepancy Qty Outbound\n" → "Discrepancy Qty Outbound"
        "Units  Received  Fast" → "Units Received Fast"
    """
    if not col_name:
        return col_name
    
    # Remove newlines and carriage returns
    normalized = col_name.replace('\n', ' ').replace('\r', ' ')
    
    # Collapse multiple spaces into single space
    normalized = ' '.join(normalized.split())
    
    # Strip leading/trailing whitespace
    normalized = normalized.strip()
    
    return normalized


def _load_xlsx_raw_values(file_content: str, sheet_name=None) -> list:
    """Load xlsx raw cell values (list of lists) via openpyxl.

    Accepts either a file path or base64-encoded content (same convention as parse_file).
    Returns [] on failure. Used by detect_source_sections and the section-aware parse_file path.
    """
    try:
        from openpyxl import load_workbook
        import base64

        is_file_path = False
        if isinstance(file_content, str) and (
            file_content.startswith("/") or file_content.startswith("C:") or
            file_content.startswith("c:") or "\\" in file_content
        ):
            is_file_path = True

        if is_file_path:
            wb = load_workbook(file_content, data_only=True, read_only=True)
        else:
            decoded_bytes = base64.b64decode(file_content)
            wb = load_workbook(io.BytesIO(decoded_bytes), data_only=True, read_only=True)

        if sheet_name is None or sheet_name == 0:
            ws = wb[wb.sheetnames[0]]
        elif isinstance(sheet_name, int):
            ws = wb[wb.sheetnames[sheet_name]]
        else:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        raw_values = []
        for row in ws.iter_rows(values_only=True):
            raw_values.append(list(row))
        return raw_values
    except Exception as e:
        print(f"   _load_xlsx_raw_values failed: {e}")
        return []


def detect_source_sections(file_content: str, file_type: str = "xlsx", sheet_name=None,
                           include_single: bool = False) -> Dict[str, Any]:
    """Detect multiple stacked sections (title + header + data block) inside a source file.

    Returns { success, sections: [...], total_rows }.
    Only xlsx/xls return sections; csv/json always return an empty list.
    Each section carries enough metadata (header_row, data_start/end, headers, row_count)
    for parse_file to later slice that single section out cleanly.

    ``include_single`` (default False): when True, single-section sheets ALSO
    return their one detected section. _detect_sections normally suppresses
    single-section detections (via the ``len(sections) >= 2`` filter at line
    ~1727) because target-side detection should treat single-section sheets
    as flat layouts. But for SOURCE-side cross-tab × per-section routing,
    a tab with one explicit titled section (e.g. April_Data has only Inbound
    Metrics — Outbound is missing) is still structurally meaningful — it
    should route into target.Inbound, not be treated as flat-and-orphaned.
    Pass include_single=True from the cross-tab × section planner; leave
    False for legacy callers.
    """
    try:
        ft = (file_type or '').lower()
        if ft not in ("xlsx", "xls", "excel"):
            return {"success": True, "sections": [], "total_rows": 0}

        raw_values = _load_xlsx_raw_values(file_content, sheet_name)
        if not raw_values:
            return {"success": True, "sections": [], "total_rows": 0}

        detected = _detect_sections(raw_values)
        # Bypass _detect_sections' ">= 2 sections" filter when caller wants
        # single sections too. We re-run the loop body without the filter.
        if include_single and not detected:
            detected = _detect_sections_single_pass(raw_values)
        sections = []
        for s in detected:
            hdrs = [str(h).strip() if h is not None else '' for h in s.get('headers', [])]
            data_start = s.get('data_start', 0)
            data_end = s.get('data_end', 0)
            row_count = max(0, data_end - data_start)
            preview_rows = []
            for r in raw_values[data_start:min(data_start + 2, data_end)]:
                preview_rows.append([str(c) if c is not None else '' for c in r])
            sections.append({
                "title": s.get('title', ''),
                "title_row": s.get('title_row', 0),
                "header_row": s.get('header_row', 0),
                "data_start": data_start,
                "data_end": data_end,
                "headers": hdrs,
                "row_count": row_count,
                "preview_rows": preview_rows,
            })

        return {
            "success": True,
            "sections": sections,
            "total_rows": len(raw_values),
        }
    except Exception as e:
        return {"success": False, "error": f"detect_source_sections failed: {str(e)}"}


def _norm_header_str(h) -> str:
    """Whitespace-and-case-normalized header for cross-sheet / cross-target compares.

    Kept local to avoid pulling in the dynamic-mapping-agent helper — same semantics,
    same casing rules. See dynamic mapping agent/lambda_function.py::_norm_header.
    """
    return ' '.join(str(h).strip().split()).lower() if h is not None else ''


def detect_source_sheets(
    file_content: str,
    file_type: str = "xlsx",
    target_headers: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Enumerate sheets in a source xlsx file and score each by target-header overlap.

    Returned shape:
    { success, sheets: [{name, headers, data_rows, score, preview_rows}], sheet_count }

    Non-xlsx file types return an empty list (with success=True) so callers can
    short-circuit the multi-sheet picker. Score is the count of source headers
    whose normalized form appears in ``target_headers``. Used by the dynamic
    mapping agent's Step 0a to decide between auto-pick and surfacing a picker.
    """
    try:
        ft = (file_type or '').lower()
        if ft not in ("xlsx", "xls", "excel"):
            return {"success": True, "sheets": [], "sheet_count": 0}

        from openpyxl import load_workbook
        import base64

        is_file_path = False
        if isinstance(file_content, str) and (
            file_content.startswith("/") or file_content.startswith("C:") or
            file_content.startswith("c:") or "\\" in file_content
        ):
            is_file_path = True

        if is_file_path:
            wb = load_workbook(file_content, data_only=True, read_only=True)
        else:
            decoded_bytes = base64.b64decode(file_content)
            wb = load_workbook(io.BytesIO(decoded_bytes), data_only=True, read_only=True)

        target_set = {_norm_header_str(h) for h in (target_headers or []) if h}

        sheets_out = []
        for name in wb.sheetnames:
            ws = wb[name]
            raw_values = []
            for row in ws.iter_rows(values_only=True):
                raw_values.append(list(row))

            # Pick the first row that *plausibly* looks like a column-header
            # row (>=2 non-empty cells AND at least one alphabetic value).
            # This skips merged-cell title banners like "Inbound Metrics"
            # that span multiple columns but collapse to a single non-empty
            # value when read via openpyxl. Without this guard the title
            # row was being picked as the header, every header column
            # except column 0 came back blank, the target-overlap score
            # was always 0, and the multi-sheet picker was forced even
            # when the underlying data shape matched the target.
            #
            # ``fallback_idx`` preserves the previous "first non-empty
            # row wins" behavior for sheets that don't have a banner —
            # e.g., a single-line metadata row with one populated cell —
            # so we never regress to "no headers detected" for legitimate
            # narrow sheets.
            header_row_idx = None
            fallback_idx = None
            for idx, row in enumerate(raw_values):
                non_empty = [c for c in row if c is not None and str(c).strip()]
                if not non_empty:
                    continue
                if fallback_idx is None:
                    fallback_idx = idx
                if (
                    len(non_empty) >= 2
                    and any(_looks_like_header_value(c) for c in non_empty)
                ):
                    header_row_idx = idx
                    break

            if header_row_idx is None:
                header_row_idx = fallback_idx

            if header_row_idx is None:
                sheets_out.append({
                    "name": name,
                    "headers": [],
                    "data_rows": 0,
                    "score": 0,
                    "meaningful_headers": 0,
                    "preview_rows": [],
                })
                continue

            headers_raw = raw_values[header_row_idx]
            headers = [str(h).strip() if h is not None else '' for h in headers_raw]

            data_rows = 0
            for row in raw_values[header_row_idx + 1:]:
                ne = [c for c in row if c is not None and str(c).strip()]
                if ne:
                    data_rows += 1

            headers_norm = {_norm_header_str(h) for h in headers if h}
            score = len(headers_norm & target_set) if target_set else 0
            meaningful = sum(1 for h in headers if h and len(h.strip()) >= 3)

            preview_rows = []
            for r in raw_values[header_row_idx + 1:header_row_idx + 3]:
                preview_rows.append([str(c) if c is not None else '' for c in r])

            sheets_out.append({
                "name": name,
                "headers": headers,
                "data_rows": data_rows,
                "score": score,
                "meaningful_headers": meaningful,
                "preview_rows": preview_rows,
            })

        return {
            "success": True,
            "sheets": sheets_out,
            "sheet_count": len(sheets_out),
        }
    except Exception as e:
        return {"success": False, "error": f"detect_source_sheets failed: {str(e)}"}


def detect_target_tab_overlap(
    tabs: List[Dict[str, Any]],
    anchor_target_col: str,
    source_anchor_values: Optional[List[Any]] = None,
    sample_limit: int = 5,
) -> Dict[str, Any]:
    """For each target tab, count how many of ``source_anchor_values`` already
    exist in the tab's anchor column. Used by the dynamic mapping agent to
    decide whether to surface a target-tab picker (Option D logic).

    Args:
        tabs: list of pre-fetched tab dicts. Each must contain at least:
              - ``name``: tab name (str)
              - ``raw_values``: 2D list of cell values
              - ``header_row_count`` (optional, defaults to 1)
        anchor_target_col: the target-side header that find_identifier picked
            as the row anchor. The function looks for this header (case- and
            whitespace-insensitive) in each tab and uses that column's values
            for overlap counting.
        source_anchor_values: anchor values pulled from the source. Empty
            iterable -> every tab returns ``overlap_count=0`` (the picker
            won't fire). Values are stringified and normalized for
            comparison (strip + lowercase + collapsed whitespace).
        sample_limit: how many matching anchor values to echo back per tab
            (used for the "matching SKU-001, SKU-002, ..." sample line in
            the picker UI).

    Returns:
        {
            "success": True,
            "tabs": [
                {
                    "name": str,
                    "anchor_column_resolved": str | None,
                    "anchor_column_index": int | None,
                    "overlap_count": int,
                    "row_count": int,
                    "sample_overlap_values": [str, ...],
                },
                ...
            ],
            "anchor_target_col": str,
            "source_anchor_value_count": int,
        }

    Errors return ``{"success": False, "error": str}``. Per-tab errors are
    swallowed (we still emit a row with ``overlap_count=0``) so a single
    flaky tab doesn't break the picker.
    """
    try:
        if not isinstance(tabs, list):
            return {"success": False, "error": "tabs must be a list of tab dicts"}
        if not anchor_target_col or not str(anchor_target_col).strip():
            return {"success": False, "error": "anchor_target_col is required"}

        target_norm = _norm_header_str(anchor_target_col)
        src_values_raw = list(source_anchor_values or [])

        def _canon_anchor(v):
            # Strip + lowercase comparison. We deliberately do NOT normalize
            # dates here — Google Sheets values come back as whatever the
            # cell's display form is, and the source is a fresh parse with
            # the same display form, so direct string compare is reliable
            # for the common case. Cross-format matching (e.g. ISO vs
            # US-locale) would belong in a heavier value normalizer; for
            # now the picker is informational and a missed match just means
            # the picker doesn't fire on that tab.
            if v is None:
                return None
            s = str(v).strip().lower()
            return s if s else None

        canon_src_set = set()
        for v in src_values_raw:
            c = _canon_anchor(v)
            if c is not None:
                canon_src_set.add(c)

        out = []
        for tab in tabs:
            try:
                name = tab.get('name') if isinstance(tab, dict) else None
                raw = tab.get('raw_values') if isinstance(tab, dict) else None
                hrc = int(tab.get('header_row_count') or 1) if isinstance(tab, dict) else 1
                if not isinstance(raw, list) or not raw:
                    out.append({
                        "name": name or "",
                        "anchor_column_resolved": None,
                        "anchor_column_index": None,
                        "overlap_count": 0,
                        "row_count": 0,
                        "sample_overlap_values": [],
                    })
                    continue

                # Anchor-column resolution: check every row that could be a
                # header (covers grouped headers up to hrc rows).
                col_idx = None
                resolved_header = None
                header_rows = raw[:max(hrc, 1)]
                for r in header_rows:
                    if not isinstance(r, list):
                        continue
                    for i, h in enumerate(r):
                        if h is None:
                            continue
                        if _norm_header_str(str(h)) == target_norm:
                            col_idx = i
                            resolved_header = str(h).strip()
                            break
                    if col_idx is not None:
                        break

                data_rows = raw[hrc:] if len(raw) > hrc else []

                if col_idx is None or not canon_src_set:
                    out.append({
                        "name": name or "",
                        "anchor_column_resolved": resolved_header,
                        "anchor_column_index": col_idx,
                        "overlap_count": 0,
                        "row_count": len(data_rows),
                        "sample_overlap_values": [],
                    })
                    continue

                overlap_count = 0
                samples = []
                seen_in_samples = set()
                for r in data_rows:
                    if not isinstance(r, list) or col_idx >= len(r):
                        continue
                    val = r[col_idx]
                    canon = _canon_anchor(val)
                    if canon is None or canon not in canon_src_set:
                        continue
                    overlap_count += 1
                    if len(samples) < sample_limit and canon not in seen_in_samples:
                        samples.append(str(val).strip() if val is not None else '')
                        seen_in_samples.add(canon)

                out.append({
                    "name": name or "",
                    "anchor_column_resolved": resolved_header,
                    "anchor_column_index": col_idx,
                    "overlap_count": overlap_count,
                    "row_count": len(data_rows),
                    "sample_overlap_values": samples,
                })
            except Exception as inner_err:
                out.append({
                    "name": (tab.get('name') if isinstance(tab, dict) else '') or '',
                    "anchor_column_resolved": None,
                    "anchor_column_index": None,
                    "overlap_count": 0,
                    "row_count": 0,
                    "sample_overlap_values": [],
                    "error": f"tab scan failed: {inner_err}",
                })

        return {
            "success": True,
            "tabs": out,
            "anchor_target_col": anchor_target_col,
            "source_anchor_value_count": len(canon_src_set),
        }
    except Exception as e:
        return {"success": False, "error": f"detect_target_tab_overlap failed: {str(e)}"}


@monitor_task("mapping_agent", "parse_file")
def parse_file(file_content: str, file_type: str = "csv", sheet_name: str = None, section: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Parse uploaded file content OR file path into structured data

    Args:
        file_content: File content as string/bytes OR file path
        file_type: Type of file (csv, xlsx, xls, excel, json)
        sheet_name: Optional sheet name for Excel files (default: first sheet)

    Returns:
        Dictionary with parsed data, columns, and metadata
    """
    try:
        # Check if file_content is actually a file path
        is_file_path = False
        if isinstance(file_content, str) and (
            file_content.startswith("/") or file_content.startswith("C:") or 
            file_content.startswith("c:") or "\\" in file_content
        ):
            is_file_path = True
            print(f"Detected file path: {file_content}")

        # Parse based on file type
        if file_type.lower() == "csv":
            if is_file_path:
                df = pd.read_csv(file_content)
            else:
                # FIX: Decode Base64 CSV content first
                try:
                    import base64
                    print(f"   Decoding base64 CSV file...")
                    decoded_text = base64.b64decode(file_content).decode('utf-8')
                    print(f"   Decoded {len(decoded_text)} characters")
                    df = pd.read_csv(io.StringIO(decoded_text))
                    print(f"   Successfully parsed CSV file")
                except Exception as csv_error:
                    print(f"   CSV parsing error: {str(csv_error)}")
                    return {
                        "success": False,
                        "error": f"Failed to parse CSV file: {str(csv_error)}",
                    }

        elif file_type.lower() in ["xlsx", "xls", "excel"]:
            if section:
                try:
                    print(f"   Slicing section: title='{section.get('title','')}' header_row={section.get('header_row')} data=[{section.get('data_start')}:{section.get('data_end')}]")
                    raw_values = _load_xlsx_raw_values(file_content, sheet_name)
                    if not raw_values:
                        return {"success": False, "error": "Failed to read raw xlsx values for section slicing"}

                    header_row = int(section.get('header_row', 0))
                    data_start = int(section.get('data_start', header_row + 1))
                    data_end   = int(section.get('data_end', len(raw_values)))

                    if header_row >= len(raw_values):
                        return {"success": False, "error": f"section.header_row {header_row} out of range ({len(raw_values)} rows)"}

                    raw_headers = raw_values[header_row]
                    headers = []
                    for i, h in enumerate(raw_headers):
                        if h is None or str(h).strip() == "":
                            headers.append(f"Column_{i}")
                        else:
                            headers.append(str(h).strip())

                    data_rows = raw_values[data_start:data_end]
                    max_cols = len(headers)
                    padded = []
                    for r in data_rows:
                        row_list = list(r) if r is not None else []
                        if len(row_list) < max_cols:
                            row_list = row_list + [None] * (max_cols - len(row_list))
                        elif len(row_list) > max_cols:
                            row_list = row_list[:max_cols]
                        padded.append(row_list)

                    df = pd.DataFrame(padded, columns=headers)
                    print(f"   Built sectioned DataFrame: {df.shape[0]} rows x {df.shape[1]} cols")
                except Exception as e:
                    print(f"   Section slicing error: {str(e)}")
                    return {"success": False, "error": f"Failed to slice section: {str(e)}"}
            elif is_file_path:
                df = pd.read_excel(file_content, sheet_name=sheet_name)
            else:
                try:
                    import base64
                    print(f"   Decoding base64 Excel file...")
                    decoded_bytes = base64.b64decode(file_content)
                    print(f"   Decoded {len(decoded_bytes)} bytes")
                    
                    df = pd.read_excel(io.BytesIO(decoded_bytes), sheet_name=sheet_name)
                    print(f"   Successfully parsed Excel file")

                except Exception as e:
                    print(f"   Excel parsing error: {str(e)}")
                    return {
                        "success": False,
                        "error": f"Failed to parse Excel file: {str(e)}. File content type: {type(file_content)}, length: {len(file_content) if isinstance(file_content, str) else 'N/A'}",
                    }

        elif file_type.lower() == "json":
            if is_file_path:
                df = pd.read_json(file_content)
            else:
                df = pd.read_json(io.StringIO(file_content))
        else:
            return {
                "success": False,
                "error": f"Unsupported file type: {file_type}. Supported: csv, xlsx, xls, excel, json",
            }

        # Clean the data
        df = df.dropna(how="all")
        df.columns = df.columns.astype(str)

        # FIX: Remove duplicate "print"
        print(f"   Normalizing {len(df.columns)} column names...")
        original_cols = df.columns.tolist()
        df.columns = [normalize_column_name(col) for col in df.columns]
        
        # Show which columns were changed
        changes_count = sum(1 for orig, norm in zip(original_cols, df.columns) if orig != norm)
        if changes_count > 0:
            print(f"   Normalized {changes_count} column names")
            for orig, norm in zip(original_cols, df.columns):
                if orig != norm:
                    print(f"      '{orig}' → '{norm}'")

        # Get metadata
        columns = df.columns.tolist()
        row_count = len(df)

        # Get sample data (first 5 rows)
        sample_df = df.head(5)

        # Infer data types for each column
        data_types = {}
        for col in columns:
            dtype = str(df[col].dtype)
            if "int" in dtype:
                data_types[col] = "integer"
            elif "float" in dtype:
                data_types[col] = "float"
            elif "datetime" in dtype:
                data_types[col] = "datetime"
            elif "bool" in dtype:
                data_types[col] = "boolean"
            else:
                data_types[col] = "string"

        # Get sample values for each column
        sample_values = {}
        for col in columns:
            non_null = df[col].dropna().head(3).tolist()
            sample_values[col] = [str(val) for val in non_null]

        return {
            "success": True,
            "columns": columns,
            "row_count": row_count,
            "column_count": len(columns),
            "data_types": data_types,
            "sample_values": sample_values,
            "sample_data": json.loads(sample_df.to_json(orient="records", date_format="iso")),
            "full_data": df.to_json(orient="records"),
            "metadata": {
                "parsed_at": datetime.now().isoformat(),
                "file_type": file_type,
                "sheet_name": sheet_name,
                "has_header": True,
                "encoding": "utf-8",
            },
        }

    except Exception as e:
        return {"success": False, "error": f"Failed to parse file: {str(e)}"}


@monitor_task("mapping_agent", "smart_column_mapping")
def smart_column_mapping(
    source_columns: Any = None,
    target_columns: List[str] = None,
    sample_data: Optional[List[Dict]] = None,
    source_data_types: Optional[Dict[str, str]] = None,
    sample_values: Optional[Dict[str, List[str]]] = None,
    skip_temporal: bool = True,
    skip_calculated: bool = True,
    data: Any = None,
) -> Dict[str, Any]:
    """
    Intelligently map source columns to target columns using AI/heuristics

    Args:
        source_columns: List of source column names (or string representation)
        target_columns: List of target column names (optional, uses SAFEXPRESSOPS_TARGET_COLUMNS by default)
        sample_data: Optional sample data for better analysis
        source_data_types: Optional data types for source columns
        sample_values: Optional sample values for each source column
        skip_temporal: If True, excludes temporal columns from target
        data: DEPRECATED - Alias for source_columns (for backwards compatibility)

    Returns:
        Dictionary with mappings, confidence scores, and recommendations
    """
    try:
        # HANDLE BACKWARDS COMPATIBILITY
        if source_columns is None and data is not None:
            print(f"Received 'data' parameter instead of 'source_columns' - using as alias")
            source_columns = data

        if source_columns is None:
            return {
                "success": False,
                "error": "Missing required parameter: 'source_columns' (or 'data' for backwards compatibility)",
            }

        print(f"\nSmart Column Mapping - Input Validation")
        print(f"   source_columns type: {type(source_columns).__name__}")

        if isinstance(source_columns, str):
            print(f"   source_columns is a string, parsing...")
            print(f"   String length: {len(source_columns)}")
            print(f"   First 100 chars: {source_columns[:100]}")

            import ast

            parsed = None

            # Strategy 1: JSON loads
            try:
                import json
                parsed = json.loads(source_columns)
                print(f"   Parsed with json.loads() - {len(parsed)} items")

                # Check if parsed data is a list of dicts (full data) instead of column names
                if isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], dict):
                    print(f"   Detected full data (list of dicts) instead of column names")
                    print(f"   Extracting column names from first row...")
                    parsed = list(parsed[0].keys())
                    print(f"   Extracted {len(parsed)} column names: {parsed[:5]}...")

            except json.JSONDecodeError as e1:
                print(f"   json.loads() failed: {str(e1)}")

                # Strategy 2: Fix quotes and retry
                try:
                    fixed = source_columns.replace("'", '"')
                    parsed = json.loads(fixed)
                    print(f"   Parsed after fixing quotes - {len(parsed)} items")

                    if isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], dict):
                        print(f"   Detected full data (list of dicts) instead of column names")
                        print(f"   Extracting column names from first row...")
                        parsed = list(parsed[0].keys())
                        print(f"   Extracted {len(parsed)} column names: {parsed[:5]}...")

                except json.JSONDecodeError as e2:
                    print(f"   Quote fix failed: {str(e2)}")

                    # Strategy 3: ast.literal_eval
                    try:
                        parsed = ast.literal_eval(source_columns)
                        print(f"   Parsed with ast.literal_eval() - {len(parsed)} items")

                        if isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], dict):
                            print(f"   Detected full data (list of dicts) instead of column names")
                            print(f"   Extracting column names from first row...")
                            parsed = list(parsed[0].keys())
                            print(f"   Extracted {len(parsed)} column names: {parsed[:5]}...")

                    except (ValueError, SyntaxError) as e3:
                        return {
                            "success": False,
                            "error": f"Could not parse source_columns: {str(e3)}",
                        }

            source_columns = parsed

        # Validate it's now a list
        if not isinstance(source_columns, list):
            return {
                "success": False,
                "error": f"source_columns must be a list, got {type(source_columns).__name__}. Value: {str(source_columns)[:200]}",
            }

        if len(source_columns) == 0:
            return {"success": False, "error": "source_columns is empty"}

        # Validate all elements are strings
        non_string_items = [item for item in source_columns if not isinstance(item, str)]
        if non_string_items:
            return {
                "success": False,
                "error": f"source_columns must be a list of strings (column names), but contains {len(non_string_items)} non-string items. First non-string: {type(non_string_items[0]).__name__} = {str(non_string_items[0])[:100]}. Did you pass full_data instead of columns?",
            }

        print(f"   Validated source_columns: {len(source_columns)} columns")
        print(f"   First 5 columns: {source_columns[:5]}")

        from safexpressops_target_columns import (
            SAFEXPRESSOPS_OPERATIONAL_ONLY,
            TEMPORAL_COLUMNS,
            INPUT_COLUMNS,
            CALCULATED_COLUMNS,
        )

        # Filter calculated columns from SOURCE
        if skip_calculated:
            def normalize_for_comparison(name: str) -> str:
                if not name:
                    return ""
                name = name.replace("\\n", " ").replace("\n", " ")
                name = " ".join(name.split())
                return name.strip().lower()

            calc_normalized = [normalize_for_comparison(c) for c in CALCULATED_COLUMNS]

            source_columns_original = len(source_columns)
            source_columns_filtered = [
                col for col in source_columns 
                if normalize_for_comparison(col) not in calc_normalized
            ]

            filtered_count = source_columns_original - len(source_columns_filtered)
            if filtered_count > 0:
                print(f"\n SOURCE Column Filtering:")
                print(f"   Original source columns:  {source_columns_original}")
                print(f"   Calculated filtered out:  {filtered_count}")
                print(f"   Remaining to map:         {len(source_columns_filtered)}")

            source_columns = source_columns_filtered

        # Use INPUT_COLUMNS if skip_calculated is True
        if target_columns is None:
            if skip_calculated:
                target_columns = INPUT_COLUMNS
                print("Using INPUT_COLUMNS (78 mappable columns)")
                print("   Excluding 36 calculated columns (formulas preserved)")
            elif skip_temporal:
                target_columns = SAFEXPRESSOPS_OPERATIONAL_ONLY
                print("Using SAFEXPRESSOPS_OPERATIONAL_ONLY (114 operational columns)")
            else:
                target_columns = SAFEXPRESSOPS_TARGET_COLUMNS
                print("Using SAFEXPRESSOPS_TARGET_COLUMNS (all 118 columns)")

        if SMART_MAPPING_AVAILABLE:
            print("Using SmartMappingEngine for AI-powered mapping...")

            sample_df = None
            if sample_data:
                try:
                    if isinstance(sample_data, str):
                        import json
                        sample_data = json.loads(sample_data)

                    if isinstance(sample_data, list):
                        sample_df = pd.DataFrame(sample_data)
                    elif isinstance(sample_data, dict):
                        sample_df = pd.DataFrame([sample_data])
                    else:
                        print(f"Unexpected sample_data type: {type(sample_data)}, ignoring")
                        sample_df = None

                    if sample_df is not None and not sample_df.empty:
                        print(f"   Sample data converted: {len(sample_df)} rows, {len(sample_df.columns)} columns")
                    else:
                        print(f"   No valid sample data provided")
                        sample_df = None

                except Exception as e:
                    print(f"Warning: Could not convert sample_data to DataFrame: {str(e)}")
                    print(f"   Continuing without sample data")
                    sample_df = None

            smart_engine = SmartMappingEngine()
            result = smart_engine.smart_map_columns(
                source_columns=source_columns,
                target_columns=target_columns,
                sample_data=sample_df,
            )

            result = filter_safe_mappings(result)

            mappings = {}
            confidence_scores = {}
            needs_review = []

            for source_col, mapping_info in result["mappings"].items():
                mappings[source_col] = mapping_info["target"]
                confidence_scores[source_col] = mapping_info["confidence_score"]

                if mapping_info["needs_review"]:
                    needs_review.append({
                        "source_column": source_col,
                        "suggested_target": mapping_info["target"],
                        "confidence": mapping_info["confidence_score"],
                        "reason": f"Low confidence ({mapping_info['confidence_level']})",
                    })

            return {
                "success": True,
                "mappings": mappings,
                "confidence_scores": confidence_scores,
                "needs_review": needs_review,
                "high_confidence_count": result["summary"]["high_confidence_mappings"],
                "accuracy_estimate": result["summary"]["accuracy_estimate"],
                "method": "smart_mapping_engine",
            }
        else:
            # Fallback: Simple string similarity matching
            print("Using fallback string similarity matching...")
            from difflib import SequenceMatcher

            mappings = {}
            confidence_scores = {}
            needs_review = []

            for source_col in source_columns:
                best_match = None
                best_score = 0.0

                for target_col in target_columns:
                    score = SequenceMatcher(None, source_col.lower(), target_col.lower()).ratio()

                    if score > best_score:
                        best_score = score
                        best_match = target_col

                mappings[source_col] = best_match if best_score > 0.3 else None
                confidence_scores[source_col] = best_score

                if best_score < 0.7:
                    needs_review.append({
                        "source_column": source_col,
                        "suggested_target": best_match,
                        "confidence": best_score,
                        "reason": "Low string similarity",
                    })

            high_confidence = sum(1 for score in confidence_scores.values() if score >= 0.7)

            return {
                "success": True,
                "mappings": mappings,
                "confidence_scores": confidence_scores,
                "needs_review": needs_review,
                "high_confidence_count": high_confidence,
                "accuracy_estimate": (sum(confidence_scores.values()) / len(confidence_scores) if confidence_scores else 0),
                "method": "string_similarity_fallback",
            }
    except Exception as e:
        print(f"Smart mapping error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": f"Mapping failed: {str(e)}"}


def filter_safe_mappings(mappings_result: Dict[str, Any]) -> Dict[str, Any]:
    """Filter out unsafe mappings (calculated columns and incorrect matches)"""
    from safexpressops_target_columns import is_calculated_column

    print("\n Filtering unsafe mappings...")

    filtered_mappings = {}
    stats = {
        "calculated_filtered": 0,
        "incorrect_filtered": 0,
        "safe_passed": 0,
    }

    incorrect_pairs = [
        ("WH QA Incident", "Losttime Incident"),
        ("Space Utilization", "Truck Utilization"),
    ]

    for source, mapping_info in mappings_result["mappings"].items():
        target = mapping_info.get("target")

        if target is None:
            filtered_mappings[source] = mapping_info
            continue

        if is_calculated_column(target):
            print(f"   FILTERED: {source} → {target} (target has formula)")
            mapping_info["target"] = None
            mapping_info["skip_reason"] = "Target column has formula"
            mapping_info["confidence_level"] = "blocked"
            stats["calculated_filtered"] += 1
            filtered_mappings[source] = mapping_info
            continue

        if (source, target) in incorrect_pairs:
            print(f"   FILTERED: {source} → {target} (incorrect semantic match)")
            mapping_info["target"] = None
            mapping_info["skip_reason"] = "Incorrect semantic mapping"
            mapping_info["confidence_level"] = "blocked"
            stats["incorrect_filtered"] += 1
            filtered_mappings[source] = mapping_info
            continue

        print(f"   SAFE: {source} → {target}")
        stats["safe_passed"] += 1
        filtered_mappings[source] = mapping_info

    mappings_result["mappings"] = filtered_mappings
    mappings_result["summary"]["calculated_filtered"] = stats["calculated_filtered"]
    mappings_result["summary"]["incorrect_filtered"] = stats["incorrect_filtered"]
    mappings_result["summary"]["safe_mappings"] = stats["safe_passed"]

    print(f"\n   Filtering Summary:")
    print(f"      Safe mappings:        {stats['safe_passed']}")
    print(f"      Calculated filtered:  {stats['calculated_filtered']}")
    print(f"      Incorrect filtered:   {stats['incorrect_filtered']}")

    return mappings_result


def extract_dates_from_all_rows(
    data: str, date_column_name: str = "Date"
) -> Dict[str, Any]:
    """
    Extract dates from ALL rows for date-based row matching

    Args:
        data: JSON string of full data
        date_column_name: Name of the date column

    Returns:
        List of {row_index, date, data} for each row
    """
    try:
        print(f"\nExtracting dates from all rows...")
        print(f"   Date column to find: '{date_column_name}'")
        
        # Parse the JSON data
        if isinstance(data, str):
            df = pd.read_json(io.StringIO(data))
        elif isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = data
        
        print(f"   DataFrame shape: {df. shape}")
        print(f"   Available columns: {list(df.columns)[:10]}...")  # Show first 10
        
        # FIX: Case-insensitive column matching
        actual_date_column = None
        for col in df.columns:
            if str(col).strip().lower() == date_column_name. strip().lower():
                actual_date_column = col
                print(f"   Found exact match: '{col}'")
                break
        
        # Try partial match if exact match fails
        if actual_date_column is None:
            for col in df.columns:
                if date_column_name. lower() in str(col).lower():
                    actual_date_column = col
                    print(f"   Using partial match: '{col}'")
                    break
        
        if actual_date_column is None:
            print(f"   Date column not found!")
            print(f"   Available columns: {list(df.columns)}")
            return {
                "success": False,
                "error": f"Date column '{date_column_name}' not found.  Available columns: {list(df.columns)[:20]}",
                "rows_with_dates": []
            }
        
        # Sample the date values for debugging
        sample_dates = df[actual_date_column].head(5). tolist()
        print(f"   Sample date values: {sample_dates}")
        print(f"   Sample date types: {[type(d).__name__ for d in sample_dates]}")
        
        rows_with_dates = []
        failed_rows = []
        
        # IMPROVED: More comprehensive date formats
        date_formats = [
            # Excel-style formats (PRIORITY for your case: "01-Jan-25")
            "%d-%b-%y",      # 01-Jan-25
            "%d-%b-%Y",      # 01-Jan-2025
            "%d/%b/%y",      # 01/Jan/25
            "%d/%b/%Y",      # 01/Jan/2025
            # ISO formats
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            # US formats
            "%m/%d/%Y",
            "%m/%d/%y",
            # European formats
            "%d/%m/%Y",
            "%d/%m/%y",
            # Dash formats
            "%d-%m-%Y",
            "%d-%m-%y",
            # Written formats
            "%d-%B-%y",      # 01-January-25
            "%d-%B-%Y",      # 01-January-2025
            "%B %d, %Y",     # January 01, 2025
            "%b %d, %Y",     # Jan 01, 2025
        ]

        for idx, row in df. iterrows():
            date_value = row[actual_date_column]
            
            # Skip empty/null dates
            if pd. isna(date_value) or date_value == '' or date_value is None:
                failed_rows.append({
                    'row': int(idx),
                    'value': str(date_value),
                    'reason': 'Empty or null'
                })
                continue

            parsed_date = None
            
            # Case 1: Already a datetime/Timestamp
            if isinstance(date_value, (pd.Timestamp, datetime)):
                parsed_date = date_value
                if isinstance(parsed_date, pd. Timestamp):
                    parsed_date = parsed_date.to_pydatetime()
            
            # Case 2: Numeric (Excel serial date)
            elif isinstance(date_value, (int, float)) and not pd.isna(date_value):
                try:
                    # Excel serial date (days since 1899-12-30)
                    from datetime import timedelta
                    base_date = datetime(1899, 12, 30)
                    parsed_date = base_date + timedelta(days=int(date_value))
                    print(f"   Row {idx}: Parsed serial date {date_value} → {parsed_date. strftime('%Y-%m-%d')}")
                except Exception as e:
                    print(f"   Row {idx}: Failed to parse serial date {date_value}: {e}")
            
            # Case 3: String - try multiple formats
            elif isinstance(date_value, str):
                date_str = str(date_value). strip()
                
                for fmt in date_formats:
                    try:
                        parsed_date = datetime. strptime(date_str, fmt)
                        
                        # FIX: Handle 2-digit years (25 → 2025, not 1925)
                        if parsed_date. year < 100:
                            parsed_date = parsed_date.replace(year=parsed_date.year + 2000)
                        elif parsed_date.year < 1950:
                            # If year is like 1925, it was probably meant to be 2025
                            parsed_date = parsed_date.replace(year=parsed_date.year + 100)
                        
                        break
                    except ValueError:
                        continue
                
                # Fallback: Use pandas parser
                if not parsed_date:
                    try:
                        pd_parsed = pd. to_datetime(date_str, dayfirst=True)
                        if isinstance(pd_parsed, pd.Timestamp):
                            parsed_date = pd_parsed.to_pydatetime()
                    except:
                        pass

            # Add to results if parsed successfully
            if parsed_date:
                # Convert row to dict, handling numpy types
                row_dict = {}
                for col, val in row.items():
                        if pd.isna(val) if not isinstance(val, (list, dict)) else False:
                            row_dict[col] = None
                        elif isinstance(val, (pd.Timestamp, datetime)):
                            row_dict[col] = val.strftime('%Y-%m-%d')
                        elif isinstance(val, (np.integer, np.int64, np.int32)):
                            row_dict[col] = int(val)
                        elif isinstance(val, (np.floating, np.float64, np.float32)):
                            row_dict[col] = float(val)
                        elif isinstance(val, np.bool_):
                            row_dict[col] = bool(val)
                        else:
                            row_dict[col] = str(val) if not isinstance(val, (str, int, float, bool, type(None))) else val
                
                rows_with_dates. append({
                    "row_index": int(idx),
                    "date": parsed_date. strftime("%Y-%m-%d"),  # ISO format for matching
                    "date_formatted": parsed_date.strftime("%d-%b-%y"),  # Display format
                    "original_value": str(date_value),
                    "row_data": row_dict,
                })
            else:
                failed_rows.append({
                    'row': int(idx),
                    'value': str(date_value),
                    'type': type(date_value).__name__,
                    'reason': 'Could not parse'
                })

        print(f"\n   Successfully extracted {len(rows_with_dates)} dates")
        
        if rows_with_dates:
            print(f"   Date range: {rows_with_dates[0]['date']} to {rows_with_dates[-1]['date']}")
            print(f"   First: {rows_with_dates[0]['original_value']} → {rows_with_dates[0]['date']}")
        
        if failed_rows:
            print(f"   Failed to parse {len(failed_rows)} rows")
            print(f"   Sample failures: {failed_rows[:3]}")

        return {
            "success": True,
            "rows_with_dates": rows_with_dates,
            "total_rows": len(rows_with_dates),
            "failed_rows": failed_rows[:10],  # Only return first 10 failures
            "date_column": actual_date_column,
        }

    except Exception as e:
        print(f"Error extracting dates: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "success": False, 
            "error": f"Failed to extract dates: {str(e)}",
            "rows_with_dates": []
        }

@monitor_task("mapping_agent", "transform_data")
def transform_data(
    source_data: str = None,
    mappings: Any = None,
    target_columns: Optional[List[str]] = None,
    fill_missing: bool = True,
    data: str = None,
) -> Dict[str, Any]:
    """
    Apply mappings and transform source data to target structure

    Args:
        source_data: JSON string of source data (from parse_file)
        mappings: Dictionary of source -> target column mappings (or string representation)
        target_columns: Optional list of target columns (for ordering)
        fill_missing: If True, fill unmapped target columns with empty values
        data: DEPRECATED - Alias for source_data (for backwards compatibility)

    Returns:
        Transformed data ready for upload to destination
    """
    try:
        print(f"\nTransform Data - Input Validation")
        print(f"   source_data: {'present' if source_data else 'MISSING'}")
        print(f"   data: {'present' if data else 'MISSING'}")
        print(f"   mappings: {'present' if mappings else 'MISSING'}")

        # HANDLE BACKWARDS COMPATIBILITY
        if source_data is None and data is not None:
            print(f"   Received 'data' parameter instead of 'source_data'")

            if isinstance(data, dict) and "mappings" in data:
                print(f"      Detected full mapping result object, extracting fields:")
                print(f"         - mappings keys: {list(data['mappings'].keys()) if isinstance(data.get('mappings'), dict) else 'invalid'}")
                print(f"         - source_data present: {'Yes' if 'source_data' in data else 'No'}")

                if mappings is None:
                    mappings = data.get("mappings")
                    print(f"         Extracted mappings from result object")

                if "source_data" in data:
                    source_data = data.get("source_data")
                    print(f"         Extracted source_data from result object")
            else:
                print(f"      Using data as source_data")
                source_data = data

        if source_data is None:
            return {
                "success": False,
                "error": "Missing required parameter: 'source_data' (or 'data'). If passing mapping result, ensure it contains 'source_data' field.",
            }

        if mappings is None:
            return {
                "success": False,
                "error": "Missing required parameter: 'mappings'. If passing full mapping result as 'data', ensure it contains 'mappings' field.",
            }

        # DEFENSIVE TYPE CHECK
        if isinstance(mappings, str):
            print(f"Warning: mappings received as string, converting to dict")
            print(f"   Original value: {mappings[:200]}...")
            import ast

            try:
                mappings = ast.literal_eval(mappings)
                print(f"   Converted successfully")
            except (ValueError, SyntaxError) as e:
                import json
                try:
                    mappings = json.loads(mappings)
                    print(f"   Converted via JSON successfully")
                except json.JSONDecodeError:
                    return {
                        "success": False,
                        "error": f"Could not parse mappings: {str(e)}",
                    }

        if not isinstance(mappings, dict):
            return {
                "success": False,
                "error": f"mappings must be a dict, got {type(mappings).__name__}. Value: {mappings}",
            }

        print(f"\nTransform Data")
        print(f"   Mappings ({len(mappings)}): {mappings}")

        # Parse source data
        source_df = pd.read_json(source_data)

        print(f"   Source data shape: {source_df.shape}")
        print(f"   Source columns: {list(source_df.columns)}")

        # Create new dataframe with target structure
        transformed_rows = []

        for _, source_row in source_df.iterrows():
            target_row = {}

            # Apply mappings
            for source_col, target_col in mappings.items():
                if target_col and source_col in source_row:
                    value = source_row[source_col]
                    if pd.notna(value):
                        target_row[target_col] = str(value).strip()
                    else:
                        target_row[target_col] = ""

            # Fill missing target columns if requested
            if fill_missing and target_columns:
                for col in target_columns:
                    if col not in target_row:
                        target_row[col] = ""

            transformed_rows.append(target_row)

        # Convert to DataFrame
        transformed_df = pd.DataFrame(transformed_rows)

        # Reorder columns if target_columns specified
        if target_columns:
            available_cols = [col for col in target_columns if col in transformed_df.columns]
            transformed_df = transformed_df[available_cols]

        print(f"   Transformed shape: {transformed_df.shape}")
        print(f"   Transformed columns: {list(transformed_df.columns)}")

        mapped_columns = [col for col in mappings.values() if col]
        unmapped_source = [col for col in source_df.columns if col not in mappings]

        return {
            "success": True,
            "transformed_data": transformed_df.to_json(orient="records"),
            "row_count": len(transformed_df),
            "column_count": len(transformed_df.columns),
            "columns": transformed_df.columns.tolist(),
            "statistics": {
                "source_columns": len(source_df.columns),
                "target_columns": len(transformed_df.columns),
                "mapped_columns": len(mapped_columns),
                "unmapped_source_columns": len(unmapped_source),
                "rows_processed": len(transformed_df),
            },
            "unmapped_source_columns": unmapped_source,
        }

    except Exception as e:
        print(f"Transformation error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": f"Transformation failed: {str(e)}"}


# ============================================================
# NEW: OPR-SPECIFIC FUNCTIONS
# ============================================================

def merge_dates_and_transformed_data(
    dates_result: Dict[str, Any],
    transform_result: Dict[str, Any]
) -> Dict[str, Any]:
    """
    NEW: Merge extract_dates_from_all_rows + transform_data results
    
    This is the CRITICAL function that OPR workflow needs!
    
    Args:
        dates_result: Output from extract_dates_from_all_rows()
        transform_result: Output from transform_data()
    
    Returns:
        Combined data with dates + transformed row data
    """
    try:
        print(f"\nMerging dates with transformed data...")
        
        # Extract data
        rows_with_dates = dates_result.get("rows_with_dates", [])
        transformed_data_json = transform_result.get("transformed_data")
        
        if not rows_with_dates:
            return {
                "success": False,
                "error": "No dates found in dates_result"
            }
        
        if not transformed_data_json:
            return {
                "success": False,
                "error": "No transformed_data found in transform_result"
            }
        
        # Parse transformed data
        transformed_data = json.loads(transformed_data_json)
        
        print(f"   Dates rows: {len(rows_with_dates)}")
        print(f"   Transformed rows: {len(transformed_data)}")
        
        # Merge
        merged = []
        for date_row, data_row in zip(rows_with_dates, transformed_data):
            merged.append({
                "date": date_row["date"],  # ISO format: "2025-01-01"
                "date_formatted": date_row["date_formatted"],  # Display: "01-Jan-25"
                "row_data": data_row  # Transformed data with target column names
            })
        
        print(f"   Merged {len(merged)} rows")
        if merged:
            print(f"   Sample: {merged[0]['date']} → {len(merged[0]['row_data'])} columns")
        
        return {
            "success": True,
            "rows_with_dates": merged,
            "total_rows": len(merged),
            "date_column": dates_result.get("date_column", "Date")
        }
        
    except Exception as e:
        print(f"Merge error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": f"Failed to merge: {str(e)}"}

def _infer_col_type(samples: list) -> str:
    from datetime import datetime
    import re

    if not samples:
        return 'unknown'

    date_hits = id_hits = numeric_hits = 0
    total = len(samples)

    for v in samples:
        s = str(v).strip()
        matched_date = False
        # Keep in sync with _normalize_date_value() in dynamic mapping agent/lambda_function.py
        for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S',
                    '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f',
                    '%Y-%m-%dT%H:%M:%S.%fZ',
                    '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d',
                    '%m-%d-%Y', '%d-%m-%Y',
                    '%d-%b-%y', '%d-%b-%Y', '%b-%d-%Y',
                    '%d %b %Y', '%d %B %Y',
                    '%B %d %Y', '%b %d %Y',
                    '%b %d, %Y', '%B %d, %Y'):
            try:
                datetime.strptime(s, fmt)
                date_hits += 1
                matched_date = True
                break
            except ValueError:
                pass
        if not matched_date:
            # Pandas to_json default emits datetimes as epoch-milliseconds; keep
            # those from degrading to 'numeric' in sample-based type inference.
            try:
                n = float(s)
                if 1e12 <= n <= 1e14:
                    date_hits += 1
                    matched_date = True
            except ValueError:
                pass
        if re.match(r'^[A-Z]{1,6}-?\d{3,}$', s) or re.match(r'^[A-Z]{2,}\d{4,}$', s):
            id_hits += 1
        try:
            float(s.replace(',', ''))
            numeric_hits += 1
        except ValueError:
            pass

    if date_hits / total > 0.5:    return 'date'
    if id_hits / total > 0.5:      return 'id'
    if numeric_hits / total > 0.5: return 'numeric'
    return 'text'

def _detect_multi_row_headers(raw_values):
    """Detect grouped/hierarchical headers spanning 2+ rows."""
    if len(raw_values) < 3:
        return None, 1, {}

    row0 = [str(c).strip() if c else '' for c in raw_values[0]]
    row1 = [str(c).strip() if c else '' for c in raw_values[1]]
    max_len = max(len(row0), len(row1))

    blanks_with_sub = 0
    for i in range(max_len):
        r0 = row0[i] if i < len(row0) else ''
        r1 = row1[i] if i < len(row1) else ''
        if not r0 and r1:
            blanks_with_sub += 1

    if blanks_with_sub < 2:
        return None, 1, {}

    # Carry the banner forward across the FULL width (max_len), not just len(row0).
    # When the target sheet is read via the Google Sheets API, trailing blank cells
    # in row 1 are truncated, so row0 may be shorter than row1. Without the
    # bounded loop the banner ("Inbound Metrics") only spreads to the cells it
    # literally occupies, producing inconsistent grouped names like
    # ['Inbound Metrics > Date', 'Trucks', 'Pallets'] — which then prevents
    # source-target verbatim matching and breaks the writer's anchor lookup.
    filled = []
    last_val = ''
    for i in range(max_len):
        c = row0[i] if i < len(row0) else ''
        if c:
            last_val = c
        filled.append(last_val)

    headers = []
    composite_map = {}
    for i in range(max_len):
        group = filled[i] if i < len(filled) else ''
        sub = row1[i] if i < len(row1) else ''
        if group and sub and group != sub:
            name = f"{group} > {sub}"
        elif sub:
            name = sub
        elif group:
            name = group
        else:
            name = f"Column_{i}"
        headers.append(name)
        composite_map[name] = i

    return headers, 2, composite_map


def _looks_like_header_value(s):
    """Return True when ``s`` plausibly looks like a column-header label.

    Heuristic: a genuine header (e.g., "Date", "Trucks", "Pallets", "Q1 Sales")
    contains at least one alphabetic character. Pure numeric or date-formatted
    values (e.g., "2025-03-02", "424", "01/05/2025") return False. Used by
    :func:`_detect_sections` and :func:`detect_source_sheets` to reject sparse
    data rows that would otherwise be mis-classified as a new section's
    header row, producing phantom sections with garbage titles (e.g., a date
    as the section title) and date-as-header chips. See the previous-bugs
    notes for TC-L06 / TC-L03 where leaky data rows like
    ``2025-03-01 | (blank) | (blank)`` followed by ``2025-03-02 | 424`` were
    detected as a phantom second section.
    """
    if s is None:
        return False
    s = str(s).strip()
    if not s:
        return False
    return any(c.isalpha() for c in s)


def _detect_sections(raw_values, min_section_rows=0):
    """Detect multiple data sections separated by blank rows or title rows.

    ``min_section_rows`` defaults to 0 so template-style multi-section
    targets (title + header row + *no data yet*) are still recognized.
    Catching these is important for the "seed an empty OPS_DASHBOARD and
    upload data" workflow — otherwise the second empty section gets
    dropped, the top-level ``>= 2 sections`` check fails, and the whole
    sheet gets mis-classified as a single grouped-header row_per_date
    target (which then silently appends at the bottom instead of writing
    into the matching section).

    FOLLOW-UP / DEBT:
        This function is a duplicate of
        ``dynamic mapping agent/lambda_function.py::_detect_sections_local``.
        Both copies must stay in lock-step. Planned follow-up: extract into
        a shared ``common/`` module uploaded as a Lambda layer. Until then,
        any behavior change here MUST be applied to the dynamic mapping
        agent copy as well.
    """
    sections = []
    i = 0
    while i < len(raw_values):
        row = raw_values[i]
        non_empty = [c for c in row if c and str(c).strip()]

        if len(non_empty) == 0:
            i += 1
            continue

        if len(non_empty) == 1:
            title = str(non_empty[0]).strip()
            if i + 1 < len(raw_values):
                header_row = raw_values[i + 1]
                header_vals = [str(c).strip() for c in header_row if c and str(c).strip()]
                # Require >=2 non-empty header cells AND at least one cell
                # that looks like a real column-name label (contains
                # alphabetic chars). This rejects sparse data rows like
                # "2025-03-02, 424" that would otherwise be mis-classified
                # as a new section's header row, producing phantom sections
                # with date-shaped titles and date-as-header chips.
                if len(header_vals) >= 2 and any(_looks_like_header_value(v) for v in header_vals):
                    # Convert None cells to '' instead of the literal
                    # string "None" so the FE's filter(Boolean) chip
                    # rendering correctly hides empties and downstream
                    # header_index keys aren't polluted.
                    sec_headers = [str(c).strip() if c is not None else '' for c in header_row]
                    sec_header_index = {h: idx for idx, h in enumerate(sec_headers) if h}
                    data_start = i + 2
                    data_end = data_start
                    while data_end < len(raw_values):
                        r = raw_values[data_end]
                        ne = [c for c in r if c and str(c).strip()]
                        if len(ne) == 0:
                            break
                        if len(ne) == 1 and data_end > data_start:
                            # A single-non-empty row mid-data is only a
                            # real section break if the NEXT row contains
                            # plausible header strings (alphabetic). When
                            # the next row is also sparse / numeric-only,
                            # this is leaky data, not a section title —
                            # keep extending the current section's range.
                            next_idx = data_end + 1
                            if next_idx < len(raw_values):
                                next_row = raw_values[next_idx]
                                next_ne = [str(c).strip() for c in next_row if c and str(c).strip()]
                                if (
                                    len(next_ne) >= 2
                                    and any(_looks_like_header_value(v) for v in next_ne)
                                ):
                                    break
                            data_end += 1
                            continue
                        data_end += 1
                    if data_end - data_start >= min_section_rows:
                        sections.append({
                            'title': title,
                            'title_row': i,
                            'header_row': i + 1,
                            'data_start': data_start,
                            'data_end': data_end,
                            'headers': sec_headers,
                            'header_index': sec_header_index,
                        })
                    # Advance past the header row even when the section
                    # had zero data rows, so we keep scanning for the
                    # next title/header pair instead of treating the
                    # header row itself as a potential title.
                    i = max(data_end, i + 2)
                    continue
        i += 1

    return sections if len(sections) >= 2 else []


def _detect_sections_single_pass(raw_values, min_section_rows=0):
    """Variant of _detect_sections that returns 1+ sections WITHOUT the
    ``>= 2 sections`` suppression at the bottom. Used by source-side
    cross-tab × per-section routing where a tab with exactly one explicit
    titled section is structurally meaningful.

    Also detects HEADERLESS sections — i.e. a section whose data starts
    with a header row but has NO preceding title row. The canonical case
    is the user's Outbound Metrics tab where:

        row 1: Date | Dispatched | Cases   (headers, no title above)
        row 2: 2025-05-01 | 3 | 60
        row 3: 2025-05-02 | 6 | 120
        row 4: (blank)
        row 5: (blank)
        row 6: Inbound Metrics              (title)
        row 7: Date | Trucks | Pallets      (headers)
        row 8-9: data

    Without headerless detection only the second (titled) section is
    found, which silently drops the first section's data from the
    cross-tab × section routing — half the source goes invisible.

    Implementation note: this is a copy-paste of the inner loop because
    _detect_sections returns the post-filter list — we'd need to refactor
    it to expose the raw list. Future work: extract the inner loop into a
    private helper that both functions wrap. Until then, KEEP THIS IN
    LOCK-STEP with _detect_sections's loop body when the detection rules
    change (header validation, sparse-row handling, etc.).
    """
    sections = []
    i = 0
    while i < len(raw_values):
        row = raw_values[i]
        non_empty = [c for c in row if c and str(c).strip()]
        if len(non_empty) == 0:
            i += 1
            continue

        # PATH 1: Title row (single non-empty cell) followed by a header
        # row — the canonical "Inbound Metrics" → "Date|Trucks|Pallets"
        # → data shape. This is the only path the legacy _detect_sections
        # supports.
        if len(non_empty) == 1:
            title = str(non_empty[0]).strip()
            if i + 1 < len(raw_values):
                header_row = raw_values[i + 1]
                header_vals = [str(c).strip() for c in header_row if c and str(c).strip()]
                if len(header_vals) >= 2 and any(_looks_like_header_value(v) for v in header_vals):
                    sec_headers = [str(c).strip() if c is not None else '' for c in header_row]
                    sec_header_index = {h: idx for idx, h in enumerate(sec_headers) if h}
                    data_start = i + 2
                    data_end = data_start
                    while data_end < len(raw_values):
                        r = raw_values[data_end]
                        ne = [c for c in r if c and str(c).strip()]
                        if len(ne) == 0:
                            break
                        if len(ne) == 1 and data_end > data_start:
                            next_idx = data_end + 1
                            if next_idx < len(raw_values):
                                next_row = raw_values[next_idx]
                                next_ne = [str(c).strip() for c in next_row if c and str(c).strip()]
                                if (
                                    len(next_ne) >= 2
                                    and any(_looks_like_header_value(v) for v in next_ne)
                                ):
                                    break
                            data_end += 1
                            continue
                        data_end += 1
                    if data_end - data_start >= min_section_rows:
                        sections.append({
                            'title': title,
                            'title_row': i,
                            'header_row': i + 1,
                            'data_start': data_start,
                            'data_end': data_end,
                            'headers': sec_headers,
                            'header_index': sec_header_index,
                        })
                    i = max(data_end, i + 2)
                    continue

        # PATH 2: Headerless section. The current row IS the header row
        # (no preceding title). Trigger conditions, ALL required:
        #   - row has >= 2 non-empty cells
        #   - at least one cell looks like a real column-name label
        #     (alphabetic chars; rejects "2025-03-02 | 424")
        #   - the NEXT row exists and looks like data (>= 1 non-empty
        #     cell, NOT a header itself — guards against a row of
        #     numeric values being mis-identified as headers)
        # Title is set to None — the planner will fall back to column-
        # overlap scoring without a title-similarity contribution.
        if len(non_empty) >= 2 and any(_looks_like_header_value(v) for v in non_empty):
            if i + 1 < len(raw_values):
                next_row = raw_values[i + 1]
                next_ne = [c for c in next_row if c and str(c).strip()]
                # Next row must have at least one non-empty cell AND
                # NOT itself look entirely header-shaped (otherwise we'd
                # be looking at a 2-row header stack, handled elsewhere).
                if next_ne and not all(
                    _looks_like_header_value(str(c).strip()) for c in next_ne
                ):
                    sec_headers = [str(c).strip() if c is not None else '' for c in row]
                    sec_header_index = {h: idx for idx, h in enumerate(sec_headers) if h}
                    data_start = i + 1
                    data_end = data_start
                    while data_end < len(raw_values):
                        r = raw_values[data_end]
                        ne = [c for c in r if c and str(c).strip()]
                        if len(ne) == 0:
                            break
                        # Same single-cell-mid-data guard as Path 1.
                        if len(ne) == 1 and data_end > data_start:
                            next_idx = data_end + 1
                            if next_idx < len(raw_values):
                                nrow = raw_values[next_idx]
                                nne = [str(c).strip() for c in nrow if c and str(c).strip()]
                                if (
                                    len(nne) >= 2
                                    and any(_looks_like_header_value(v) for v in nne)
                                ):
                                    break
                            data_end += 1
                            continue
                        data_end += 1
                    if data_end - data_start >= min_section_rows:
                        sections.append({
                            'title': None,  # headerless-only — no title row
                            'title_row': None,
                            'header_row': i,
                            'data_start': data_start,
                            'data_end': data_end,
                            'headers': sec_headers,
                            'header_index': sec_header_index,
                        })
                    i = max(data_end, i + 1)
                    continue

        i += 1
    return sections


def structure_target_data(raw_values: list, sheet_name: str = '', sample_size: int = 5) -> dict:
    try:
        # An empty target sheet is a valid state for append / seeding flows.
        # Return a success response with an empty schema so downstream logic
        # (AI identification, diff preview, route_write) can decide what to do
        # instead of hard-failing the entire preview.
        if not raw_values:
            return {
                'success':      True,
                'sheet_name':   sheet_name,
                'total_rows':   0,
                'total_cols':   0,
                'headers':      [],
                'header_index': {},
                'header_row_count': 0,
                'composite_to_col_index': {},
                'sections':     [],
                'col_samples':  {},
                'col_types':    {},
                'formula_cols': [],
                'raw_rows':     [],
                'is_empty_target': True,
            }

        # Detect multi-section layout FIRST — a stacked template with
        # title → header → data rows is almost always mis-matched by the
        # grouped-header detector (which would flatten the first
        # section's title into its header and ignore the second
        # section entirely). If a multi-section layout is found we use
        # the first section's headers as the flat ``headers`` array so
        # the rest of the preview logic keeps working, and the AI sees
        # the ``sections`` metadata and routes to ``multi_section``.
        sections = _detect_sections(raw_values)
        if sections:
            print(f"   Detected {len(sections)} sections: {[s['title'] for s in sections]}")
            first_sec = sections[0]
            headers = [str(h).strip() for h in first_sec['headers']]
            header_row_count = 1
            composite_to_col_index = {}
        else:
            grouped_headers, header_row_count, composite_to_col_index = _detect_multi_row_headers(raw_values)
            if grouped_headers:
                headers = grouped_headers
                print(f"   Detected {header_row_count}-row grouped headers: {headers[:5]}...")
            else:
                headers = [str(h).strip() for h in raw_values[0]]
                composite_to_col_index = {}

        data_rows = raw_values[header_row_count:]
        header_index = {h: i for i, h in enumerate(headers)}

        col_samples  = {}
        col_types    = {}
        formula_cols = []

        for col_name, col_idx in header_index.items():
            samples    = []
            is_formula = False

            for row in data_rows[:sample_size]:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None and str(val).strip():
                    val_str = str(val).strip()
                    if val_str.startswith('='):
                        is_formula = True
                    else:
                        samples.append(val_str)

            if is_formula:
                formula_cols.append(col_name)

            col_samples[col_name] = samples
            col_types[col_name]   = _infer_col_type(samples)

        return {
            'success':      True,
            'sheet_name':   sheet_name,
            'total_rows':   len(data_rows),
            'total_cols':   len(headers),
            'headers':      headers,
            'header_index': header_index,
            'header_row_count': header_row_count,
            'composite_to_col_index': composite_to_col_index,
            'sections':     sections,
            'col_samples':  col_samples,
            'col_types':    col_types,
            'formula_cols': formula_cols,
            'raw_rows':     raw_values
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

def structure_source_data(parse_result: dict, sample_size: int = 5) -> dict:
    try:
        import json as _json

        if not parse_result.get('success'):
            return {'success': False, 'error': parse_result.get('error', 'parse_file failed')}

        columns   = parse_result.get('columns', [])
        full_data = parse_result.get('full_data', '[]')

        if isinstance(full_data, str):
            rows = _json.loads(full_data)
        else:
            rows = full_data

        header_row_count = 1
        composite_to_col_index = {}

        # Detect grouped source headers: if pandas columns have blanks/Unnamed
        # and the first data row looks like sub-headers (all non-numeric strings)
        if rows and len(rows) > 1:
            blank_cols = sum(1 for c in columns
                            if not str(c).strip() or str(c).strip().startswith('Unnamed'))
            if blank_cols >= 2:
                first_row = rows[0]
                non_numeric = 0
                for col in columns:
                    raw = first_row.get(col)
                    if raw is None:
                        continue
                    val = str(raw).strip()
                    if val:
                        try:
                            float(val.replace(',', ''))
                        except ValueError:
                            non_numeric += 1
                if non_numeric >= len(columns) * 0.5:
                    header_row_count = 2
                    row0 = [str(c).strip() for c in columns]
                    row1 = [(str(first_row.get(c)).strip() if first_row.get(c) is not None else '') for c in columns]
                    filled = []
                    last_val = ''
                    for c in row0:
                        if c and not c.startswith('Unnamed'):
                            last_val = c
                        filled.append(last_val)
                    new_columns = []
                    for i in range(len(columns)):
                        group = filled[i]
                        sub = row1[i]
                        if group and sub and group != sub:
                            name = f"{group} > {sub}"
                        elif sub:
                            name = sub
                        elif group:
                            name = group
                        else:
                            name = f"Column_{i}"
                        new_columns.append(name)
                        composite_to_col_index[name] = i
                    columns = new_columns
                    rows = rows[1:]
                    # Rebuild every row dict so its keys match the new
                    # composite column names. Without this, downstream
                    # callers that read ``parse_result['full_data']``
                    # directly (e.g., the dynamic-mapping-agent's
                    # ``_preview_single_sheet`` and ``transform_data``)
                    # see the OLD pandas column keys ("Unnamed: 1", "Inbound
                    # Metrics", …) while the LLM mapping uses the NEW
                    # composite names ("Inbound Metrics > Date", …) and
                    # every lookup returns ``None`` — producing 0 anchor
                    # values and 0 transformed rows. We also surface the
                    # rebuilt list as a JSON string in ``full_data`` so
                    # the caller can pick this up via
                    # ``source_schema.get('full_data')`` instead of the
                    # stale ``parse_result['full_data']``.
                    renamed_rows = []
                    for row in rows:
                        if not isinstance(row, dict):
                            renamed_rows.append(row)
                            continue
                        orig_keys = list(row.keys())
                        new_row = {}
                        for new_name, orig_idx in composite_to_col_index.items():
                            if 0 <= orig_idx < len(orig_keys):
                                new_row[new_name] = row.get(orig_keys[orig_idx])
                        renamed_rows.append(new_row)
                    rows = renamed_rows
                    print(f"   Detected grouped source headers: {columns[:5]}...")
                    print(f"   Rebuilt {len(rows)} row(s) with composite keys")

        header_index = {str(col).strip(): i for i, col in enumerate(columns)}
        col_samples  = {}
        col_types    = {}

        for col_name in columns:
            col_str  = str(col_name).strip()
            samples  = []
            for row in rows[:sample_size]:
                val = row.get(col_name) if isinstance(row, dict) else None
                if val is None and isinstance(row, dict):
                    orig_cols = list(row.keys())
                    idx = composite_to_col_index.get(col_str)
                    if idx is not None and idx < len(orig_cols):
                        val = row.get(orig_cols[idx])
                if val is not None and str(val).strip():
                    samples.append(str(val).strip())
            col_samples[col_str] = samples
            col_types[col_str]   = _infer_col_type(samples)

        return {
            'success':      True,
            'total_rows':   len(rows),
            'total_cols':   len(columns),
            'headers':      [str(c).strip() for c in columns],
            'header_index': header_index,
            'header_row_count': header_row_count,
            'composite_to_col_index': composite_to_col_index,
            'col_samples':  col_samples,
            'col_types':    col_types,
            'raw_rows':     rows,
            # Grouped-header-aware ``full_data``: when the input had
            # composite headers we rebuilt every row with the new
            # composite keys above; for flat-header inputs this round-trips
            # the original ``parse_result['full_data']`` unchanged so
            # callers can always prefer ``source_schema['full_data']``
            # without branching on layout.
            'full_data':    _json.dumps(rows),
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}

def _section_title_similarity(src_title: str, tgt_title: str) -> int:
    """Fuzzy similarity score 0..100 for two section titles.

    Used by find_identifier's multi-section pre-resolver to break ties when
    the column-overlap signal is ambiguous (Bug for cross-section anchor
    collision: source picked "Outbound Metrics" → target had both "Inbound
    Metrics" AND "Outbound Metrics" with identical column shape, picker
    defaulted to the first section by index → wrote Outbound data into
    Inbound's columns).

    Scoring rules:
      - exact (case+whitespace-normalized) match → 100
      - jaccard token-overlap → 0..99
      - missing or empty input on either side → 0 (no signal, falls back
        to legacy column-overlap tie-break)

    Token tuning: drops trailing 'metrics' / 'data' / 'table' /
    'section' so "Outbound Metrics" vs "Outbound" still scores 100, NOT
    50. Without this the "metrics" token would dominate jaccard for any
    *_metrics vs *_metrics pairing and weaken the discriminating word.
    """
    if not src_title or not tgt_title:
        return 0
    s = ' '.join(str(src_title).lower().split())
    t = ' '.join(str(tgt_title).lower().split())
    if s == t:
        return 100
    _SUFFIXES = ('metrics', 'metric', 'data', 'table', 'section', 'summary')
    s_tokens = {tok for tok in s.split() if tok not in _SUFFIXES}
    t_tokens = {tok for tok in t.split() if tok not in _SUFFIXES}
    if not s_tokens or not t_tokens:
        return 0
    overlap = len(s_tokens & t_tokens)
    union = len(s_tokens | t_tokens)
    if not union:
        return 0
    sim = int(100 * overlap / union)
    # Cap below 100 so an exact-after-suffix-strip match still ranks
    # below the literal exact match above (preserves a discrimination
    # gradient: identical title beats "trimmed" identical title beats
    # partial overlap).
    return min(sim, 99)


def find_identifier(target_schema: dict, source_schema: dict, source_section_title: str = None) -> dict:
    """
    source_section_title: optional title of the SOURCE section the user picked
        (e.g. "Outbound Metrics"). When provided, the multi-section TARGET
        pre-resolver tie-breaks on title similarity so a source labeled
        "Outbound" routes to a target section also labeled "Outbound" instead
        of defaulting to the first section. ``None`` preserves the legacy
        column-overlap-only behavior.
    """
    import openai, os, json as _json

    try:
        formula_cols = set(target_schema.get('formula_cols', []))
        target_headers = target_schema['headers']
        source_headers = source_schema['headers']

        # --- Optimization D: exact-match columns locally (no AI needed) ---
        target_header_set = set(target_headers)
        target_lower_map = {}
        for h in target_headers:
            normalized = h.strip().lower().replace('_', ' ')
            target_lower_map[normalized] = h

        # Grouped-header tail index: when target headers are flat ("Date",
        # "Trucks") but source headers are grouped ("Inbound Metrics > Date",
        # "Inbound Metrics > Trucks"), the verbatim/normalized exact-match
        # below would miss every source column and dump them all on the AI
        # — which then sometimes hallucinates the source's COMPOSITE name
        # back as the target column ("anchor_column: 'Inbound Metrics > Date'"
        # for a target whose header is just "Date"). The downstream writer
        # then can't find that "column" in the target, returns 0 matched +
        # 0 unmatched, and falls back to a no-op append. Resolving this
        # tail-side BEFORE the AI is the cleanest fix because we already
        # have full evidence (the source header literally ends with the
        # target column name).
        #
        # Symmetric reverse direction (target grouped, source flat) is
        # handled by `target_tail_lower_map` below: for each target header
        # like "Inbound Metrics > Date" we also index "date" → full target
        # name so source "Date" maps to "Inbound Metrics > Date".
        #
        # Ambiguity guard: if a tail matches multiple target headers (e.g.,
        # target has both "Inbound > Date" and "Outbound > Date" with the
        # same tail "Date"), we DO NOT auto-resolve — fall through to the
        # AI / Fix-E section pre-resolver to make the call. This preserves
        # safety for genuinely ambiguous schemas and avoids silently picking
        # the wrong side.
        from collections import defaultdict
        target_tail_lower_map = defaultdict(list)
        for h in target_headers:
            tail = str(h).split(' > ')[-1].strip()
            if tail and tail != h.strip():
                tail_norm = tail.lower().replace('_', ' ')
                target_tail_lower_map[tail_norm].append(h)

        def _strip_group_prefix_norm(s):
            """Return the normalized tail of a 'Group > Sub' header, or '' for flat."""
            if not s:
                return ''
            parts = str(s).split(' > ')
            if len(parts) < 2:
                return ''
            return parts[-1].strip().lower().replace('_', ' ')

        exact_mappings = {}
        unmatched_sources = []
        for src in source_headers:
            normalized = src.strip().lower().replace('_', ' ')
            if src in target_header_set:
                exact_mappings[src] = src
            elif normalized in target_lower_map:
                exact_mappings[src] = target_lower_map[normalized]
            else:
                # Source-side tail: 'Inbound Metrics > Date' → 'Date'.
                # Try to resolve against flat target headers via
                # target_lower_map first (the common case), then against
                # other grouped target headers via the tail index.
                src_tail_norm = _strip_group_prefix_norm(src)
                if src_tail_norm and src_tail_norm in target_lower_map:
                    exact_mappings[src] = target_lower_map[src_tail_norm]
                    continue
                # Target-side tail: source 'Date' → target 'Inbound Metrics > Date'.
                # Only resolve when EXACTLY ONE grouped target header has
                # this tail (otherwise it's ambiguous; fall through to AI).
                if normalized in target_tail_lower_map and len(target_tail_lower_map[normalized]) == 1:
                    exact_mappings[src] = target_tail_lower_map[normalized][0]
                    continue
                # Both sides grouped, tail-to-tail match: 'A > Date' → 'B > Date'.
                if (
                    src_tail_norm
                    and src_tail_norm in target_tail_lower_map
                    and len(target_tail_lower_map[src_tail_norm]) == 1
                ):
                    exact_mappings[src] = target_tail_lower_map[src_tail_norm][0]
                    continue
                unmatched_sources.append(src)

        print(f"Pre-matched {len(exact_mappings)} columns exactly: {list(exact_mappings.keys())}")
        if unmatched_sources:
            print(f"{len(unmatched_sources)} columns need AI: {unmatched_sources}")

        # --- Multi-section target: pre-resolve which section the source
        # actually targets, so the AI prompt is scoped to THAT section's
        # headers (not just the first section's, which is what
        # ``structure_target_data`` puts into ``target_schema['headers']``).
        # Without this, sources whose columns belong to a non-first
        # section (e.g. an "Outbound" tab against a target where
        # "Inbound Metrics" is section #1 and "Outbound Metrics" is
        # section #2) get partial mappings: only the columns that
        # coincidentally exist in the first section get mapped, the
        # rest collapse to null, and the writer fails with "No data
        # columns to write" or silently dumps the matched columns into
        # the wrong section. Picks the section with the highest count
        # of overlapping source headers (ties go to the first section
        # so single-section semantics are unchanged). The matched
        # section's index is surfaced via ``target_section_index`` in
        # the result so the writer can pin its section_override.
        sections_meta = target_schema.get('sections', []) or []
        matched_section = None
        matched_section_idx = None
        target_headers_for_prompt = list(target_headers)
        if len(sections_meta) >= 2 and source_headers:
            src_unmatched_norm = {_norm_header_str(s) for s in unmatched_sources if s}
            src_all_norm       = {_norm_header_str(s) for s in source_headers if s}
            scored = []
            for idx, sec in enumerate(sections_meta):
                sec_norm = {_norm_header_str(h) for h in sec.get('headers', []) if h}
                if not sec_norm:
                    continue
                unmatched_overlap = len(src_unmatched_norm & sec_norm)
                total_overlap     = len(src_all_norm & sec_norm)
                # Title similarity (0..100). Computed once per candidate
                # so the sort key below stays cheap. Used as the PRIMARY
                # tie-breaker when source_section_title is known —
                # otherwise it's identically 0 across all candidates and
                # the ordering reduces to the legacy column-overlap rule.
                title_sim = _section_title_similarity(
                    source_section_title or '', sec.get('title') or ''
                )
                scored.append((unmatched_overlap, total_overlap, title_sim, idx, sec))
            # Sort key tiers (Fix N — title-similarity tie-break for
            # cross-section anchor collision):
            #   1. title_sim DESC — strongest signal when the user's
            #      picked source section has a title and the target has
            #      a same-named section. "Outbound Metrics" → "Outbound
            #      Metrics" wins regardless of column overlap, which
            #      fixes the bug where Outbound source data leaked into
            #      Inbound target columns just because both sections
            #      had the same column shape.
            #   2. unmatched_overlap DESC — legacy primary signal:
            #      sections that satisfy the source's currently-unmatched
            #      columns are preferred (these are exactly the columns
            #      the AI would otherwise have to fabricate).
            #   3. total_overlap DESC — broader column-coverage signal.
            #   4. idx ASC — tie-break to first section so single-section
            #      semantics and the no-title-info case behave identically
            #      to the pre-fix code (no behavior drift for users who
            #      do not supply / cannot supply a source title).
            scored.sort(key=lambda t: (-t[2], -t[0], -t[1], t[3]))
            if scored:
                (best_unmatched_overlap, best_total_overlap,
                 best_title_sim, best_idx, best_sec) = scored[0]
                first_score = next((t for t in scored if t[3] == 0), None)
                # Rescope gating tiers:
                #
                #   A. Title-driven rescope (NEW for Fix N): when the
                #      source has a known title AND the best candidate's
                #      title similarity is strong (>=75 — covers exact
                #      matches at 100 and trimmed-suffix matches at 99
                #      generated by `_section_title_similarity`'s
                #      suffix-strip), bypass the column-overlap
                #      strictly_better guard. We TRUST the title signal
                #      because it survives cases where column overlap is
                #      misleading (Inbound + Outbound sections both with
                #      Date/Trucks/Pallets — column overlap ties at 3,
                #      title alone says "Outbound" → pick Outbound).
                #      A single column overlap is still required so a
                #      title-match against an utterly unrelated section
                #      (zero shared columns) still falls through.
                #
                #   B. Legacy column-overlap rescope (preserved): when
                #      no title signal is available OR title similarity
                #      is weak, fall back to the original threshold
                #      (>=1 unmatched overlap OR >=2 total overlap)
                #      AND strictly-better-than-first-section gate.
                #      This preserves single-section semantics and the
                #      pre-fix default for ambiguous cases.
                threshold_passed = best_unmatched_overlap >= 1 or best_total_overlap >= 2
                strictly_better = (
                    first_score is None
                    or best_total_overlap > first_score[1]
                )
                title_strong = (
                    bool(source_section_title)
                    and best_title_sim >= 75
                    and best_total_overlap >= 1
                )
                if best_idx != 0 and (title_strong or (threshold_passed and strictly_better)):
                    matched_section = best_sec
                    matched_section_idx = best_idx
                    target_headers_for_prompt = [
                        str(h).strip() for h in best_sec.get('headers', []) if h
                    ]
                    print(
                        f"Multi-section rescope: source "
                        f"{(source_section_title or '<no-title>')!r} "
                        f"matches section #{best_idx} "
                        f"'{best_sec.get('title')}' "
                        f"(title_sim={best_title_sim}, "
                        f"unmatched_overlap={best_unmatched_overlap}, "
                        f"total_overlap={best_total_overlap}, "
                        f"first_section_total_overlap="
                        f"{first_score[1] if first_score else 0}, "
                        f"trigger={'title' if title_strong else 'columns'}); "
                        f"prompt will be scoped to {len(target_headers_for_prompt)} "
                        f"section headers"
                    )

        # --- Structural-layout short-circuit (pre-AI) ---
        # Before anything else, check if the target shape screams a specific
        # strategy: cross_tab (source values match target headers),
        # horizontal (time-period target cols), or key_value (2-col
        # label/value). These layouts are frequently misclassified as
        # row_per_date / row_per_entity by the AI because a date or
        # id-looking column happens to match. Detecting them up-front avoids
        # the wrong fallback entirely and skips the AI call.
        structural = _detect_structural_layout(target_schema, source_schema, exact_mappings)
        if structural:
            print(
                f"Structural layout detected: {structural['write_strategy']} "
                f"(anchor='{structural.get('anchor_column')}') — skipping AI"
            )
            # Merge any exact matches into the structural mappings (for
            # cross_tab, exact matches like Date→Date are still valid).
            merged = dict(exact_mappings)
            merged.update(structural.get('column_mappings') or {})
            structural['column_mappings'] = merged
            structural['success'] = True
            return structural

        # If everything matched exactly, skip the AI call entirely
        if not unmatched_sources:
            print("All columns matched — skipping AI call entirely")
            # Infer strategy from column types
            strategy, anchor, src_anchor, anchor_type = _infer_strategy_local(
                target_schema, source_schema, exact_mappings)
            reasoning = 'All source columns matched target columns exactly.'

            # Even on the fast path, cross-check for entity overlap so we don't
            # silently duplicate rows when the source is an update.
            if strategy == 'append':
                overlap_anchor = _detect_entity_overlap_anchor(
                    target_schema, source_schema, exact_mappings
                )
                if overlap_anchor:
                    tgt_col, src_col, overlap_ratio = overlap_anchor
                    print(
                        f"Fast-path override 'append' → 'row_per_entity' "
                        f"(anchor='{tgt_col}', overlap={overlap_ratio:.0%})"
                    )
                    strategy = 'row_per_entity'
                    anchor = tgt_col
                    src_anchor = src_col
                    anchor_type = 'id'
                    reasoning += (
                        f" [local-override: {overlap_ratio:.0%} entity overlap on '{tgt_col}']"
                    )

            fast_result = {
                'success': True,
                'write_strategy': strategy,
                'anchor_column': anchor,
                'source_anchor': src_anchor,
                'anchor_type': anchor_type,
                'column_mappings': exact_mappings,
                'reasoning': reasoning,
            }
            # Surface the pre-resolved section so the writer pins to it
            # via section_override (mirrors the AI path below). When the
            # source matched a non-first section but every column was
            # also exact-matchable into the first section's headers
            # (because Date/Cases/etc. happen to repeat across sections),
            # the writer would otherwise auto-pick the wrong section.
            if matched_section_idx is not None:
                fast_result['target_section_index'] = matched_section_idx
                fast_result['target_section_title'] = (
                    matched_section.get('title') if matched_section else None
                )
            return fast_result

        # --- Build target summary for AI ---
        # When all sources matched exactly, we skip AI entirely (handled above).
        # When unmatched columns exist, include samples for ALL target columns
        # so the AI can match by data patterns (critical for cross-language sheets).
        id_keywords = ('id', 'code', 'key', 'sku', 'date', 'txn', 'transaction', 'no', 'number')
        has_cross_lang = len(unmatched_sources) > len(source_headers) * 0.5
        target_summary = {}
        for col in target_headers_for_prompt:
            if col in formula_cols:
                continue
            # When we rescoped to a non-first section, col_types/col_samples
            # may not have entries for the new section's headers (those
            # caches were built against the first section in
            # structure_target_data). The .get(col, default) calls below
            # fall back gracefully — the LLM still gets the header name +
            # the source samples it needs to make the mapping.
            col_type = target_schema['col_types'].get(col, 'unknown')
            col_lower = col.lower()
            if has_cross_lang or any(kw in col_lower for kw in id_keywords):
                samples = target_schema['col_samples'].get(col, [])[:3]
                target_summary[col] = {'type': col_type, 'samples': samples}
            else:
                target_summary[col] = col_type

        source_summary = {
            col: {
                'type': source_schema['col_types'].get(col, 'unknown'),
                'samples': source_schema['col_samples'].get(col, [])[:2]
            }
            for col in unmatched_sources
        }

        # Include section info if multi-section target was detected.
        # When we've already pre-resolved to a specific section, scope
        # the note narrowly so the LLM doesn't get confused about which
        # section's headers to use (we already only listed THAT section
        # in target_summary above, but the explicit directive here makes
        # it impossible for the LLM to drift back to first-section
        # headers it half-remembers from the structural-layout context).
        sections_info = ''
        sections = target_schema.get('sections', [])
        if sections:
            if matched_section is not None:
                sections_info = (
                    f"\n\nTARGET SECTION (pre-resolved by header overlap): "
                    f"\"{matched_section.get('title')}\" — columns "
                    f"{[h for h in matched_section.get('headers', []) if h]}.\n"
                    f"All mappings MUST land in this section's columns. "
                    f"Use 'multi_section' as write_strategy and pick the anchor "
                    f"from this section's columns above."
                )
            else:
                sec_summaries = [f"  - \"{s['title']}\": columns {[h for h in s['headers'] if h]}" for s in sections]
                sections_info = f"\n\nMULTI-SECTION TARGET: This sheet has {len(sections)} sections:\n" + '\n'.join(sec_summaries) + "\nIf the source data matches ONE section, use 'multi_section' strategy."

        prompt = f"""You are analyzing two data schemas to determine how source data should be written into a target Google Sheet.

TARGET SHEET ({target_schema['total_rows']} rows, {target_schema['total_cols']} columns):
Column names with types: {_json.dumps(target_summary)}{sections_info}

SOURCE — UNMATCHED columns only ({len(unmatched_sources)} of {len(source_headers)}):
{_json.dumps(source_summary, indent=2)}

ALREADY MATCHED (exact name match, do NOT re-map these):
{_json.dumps(exact_mappings)}

FORMULA COLUMNS (must NEVER be mapped to):
{list(formula_cols)}

STEP 1 — Determine the write strategy:
- row_per_date: target has a date column that uniquely identifies each row, AND source dates OVERLAP with existing target dates
- row_per_entity: target has an ID/SKU/code/name column that uniquely identifies each row, AND source IDs OVERLAP with existing target IDs
- composite_key: NO single column uniquely identifies a row — two or more columns together form the key (e.g. Date+Store, Date+Shift, Product+Region). Use when rows have duplicates in any single column.
- key_value: target is a 2-column layout with label/metric column and value column (vertical orientation, e.g. "Total Sales | 50000")
- cross_tab: target is a matrix where row headers AND column headers identify cells (e.g. Products down rows × Stores across columns). Source may be in flat/long form (row_entity, col_entity, value) with values that match target column HEADERS.
- horizontal: the column HEADERS themselves are time periods (Jan, Feb, Q1, 2024...) AND source has the same time-period columns
- multi_section: the target has clearly separated sections (title + header + data rows, repeated). Use ONLY when the target has multiple sections and source data maps to one of them.
- append: source data contains NEW records that don't exist in the target yet. Also use when there is no clear identifier column.

IMPORTANT for choosing between row_per_entity and append:
- If source IDs/codes look like NEW entries (not updates), choose "append"
- Compare source sample values with target sample values — if they don't overlap, it's new data → append

CRITICAL date-key rule (do NOT violate):
- If a date-typed column appears in ALREADY MATCHED and the target shows date-typed samples for that column, you MUST choose "row_per_date" (not "append") and set anchor_column to that target date column. Use "append" only when there is no date/id/name anchor at all.

STEP 2 — Find the identifier:
- anchor_column: the target column name(s) that make each row unique
- Single anchor: a string like "Date"
- Composite key: a JSON list like ["Date", "Store"] (use composite_key strategy)
- For key_value: the label/metric column name
- For cross_tab: the row-header column name (first column of the matrix)
- For horizontal: the entity column (e.g. "Product") — the column that identifies each row
- For multi_section: the anchor column within the matched section
- For append: null
- source_anchor: the corresponding source column name(s) (string or list matching anchor_column)

STEP 3 — Map ONLY the unmatched source columns to target columns:
- Match by meaning, not just name
- Columns may be in DIFFERENT LANGUAGES. Use data types, sample values, and semantic meaning to match.
- If column headers span multiple rows (grouped headers like "Burger Consumed > Patty"), match by the composite name.
- For cross_tab: map the source value column to the target column headers where the values should go. If source is flat (row, col, value), map the source row-entity column to the target row-header, and note that values will be placed in the matrix.
- NEVER map to formula columns
- CRITICAL — when in doubt, return NULL. A weak match silently writes wrong data into the wrong column, which is far worse than leaving the column unmapped (the user can review the unmapped list and either (a) accept the gap, (b) rename their source column, or (c) the FE will surface it as 'No match found' so they know nothing was lost). Examples of weak matches you MUST return null for:
   * Same numeric domain but different concept: "Cases" → "Expected Receiving Qty" is WRONG even though both are quantities. Cases is a unit count; "Expected Receiving Qty" is a planned/forecasted figure. Return null.
   * Keyword overlap without semantic alignment: source "Notes" → target "Reasoning" or "Comments" is acceptable; source "Notes" → target "Note Receivable" (an accounting term) is WRONG. Return null.
   * Generic/ambiguous source headers: source "Value" or "Amount" with no target column whose name strongly implies the same concept. Return null rather than guessing.
   * Cross-language matches where the data samples don't reinforce the meaning. If samples are numeric and could be anything, return null.
- ONLY return a non-null mapping when you can articulate, in one sentence, why the source column's data BELONGS in the target column (not just "they're both numeric" or "they both contain X keyword"). If you can't write that sentence, the answer is null.

RULES:
- anchor_column must be EXACTLY from: {target_headers_for_prompt} (string or list of strings) or null
- source_anchor must be EXACTLY from: {source_headers} (string or list of strings) or null
- column_mappings keys: ONLY the unmatched source columns
- column_mappings values must be EXACTLY from: {target_headers_for_prompt} or null

Return ONLY this JSON:
{{
    "write_strategy": "row_per_date|row_per_entity|composite_key|key_value|cross_tab|horizontal|multi_section|append",
    "anchor_column": "<exact target col(s): string or list or null>",
    "source_anchor": "<exact source col(s): string or list or null>",
    "anchor_type": "date|id|name|composite|null",
    "column_mappings": {{
        "<unmatched source col>": "<target col or null>"
    }},
    "reasoning": "<one sentence>"
}}"""

        client = openai.OpenAI(api_key=os.environ.get('OPENAI_API_KEY'))
        model = 'gpt-4o' if has_cross_lang else 'gpt-4o-mini'
        print(f"Calling {model} for {len(unmatched_sources)} unmatched columns (cross_lang={has_cross_lang})...")
        resp = client.chat.completions.create(
            model=model,
            messages=[{'role': 'user', 'content': prompt}],
            temperature=0,
            response_format={"type": "json_object"}
        )

        result = _json.loads(resp.choices[0].message.content)
        # Merge exact matches with AI matches
        ai_mappings = result.get('column_mappings', {})
        merged = {**exact_mappings, **ai_mappings}
        result['column_mappings'] = merged
        result['success'] = True

        # Post-AI repair pass: even with the pre-AI grouped-header tail
        # match above, the AI sometimes still returns a SOURCE composite
        # name as a target column (e.g., 'anchor_column': 'Inbound Metrics
        # > Date' for a target with flat header 'Date'). The downstream
        # writer can't find that "column" in the target sheet and silently
        # no-ops. Detect any value that isn't a valid target header and
        # try to repair it by stripping the group prefix and re-resolving
        # against target_lower_map / target_tail_lower_map (same logic as
        # the pre-AI exact-match phase). Repair touches anchor_column,
        # source_anchor (when it borrowed a wrong target name), and every
        # column_mappings value. Anything that still can't be resolved is
        # set to null so the validation guards downstream surface a clear
        # error rather than writing into the void.
        valid_target_headers = set(target_headers_for_prompt) | set(target_headers)

        def _repair_target_name(val):
            """Coerce val to a valid target header or None.

            Tries (in order):
              1. val is already a valid target header → keep
              2. val's tail (after ' > ') matches target_lower_map → use full target name
              3. val matches target_tail_lower_map (uniquely) → use full target name
              4. fall through → None (caller decides whether to drop or error)
            """
            if val is None or val == '':
                return None
            if val in valid_target_headers:
                return val
            v_norm = str(val).strip().lower().replace('_', ' ')
            if v_norm in target_lower_map:
                return target_lower_map[v_norm]
            v_tail_norm = _strip_group_prefix_norm(val)
            if v_tail_norm and v_tail_norm in target_lower_map:
                return target_lower_map[v_tail_norm]
            if v_norm in target_tail_lower_map and len(target_tail_lower_map[v_norm]) == 1:
                return target_tail_lower_map[v_norm][0]
            if (
                v_tail_norm
                and v_tail_norm in target_tail_lower_map
                and len(target_tail_lower_map[v_tail_norm]) == 1
            ):
                return target_tail_lower_map[v_tail_norm][0]
            return None

        # Repair anchor_column (single string OR list-of-strings for composites).
        anc = result.get('anchor_column')
        if isinstance(anc, str):
            repaired = _repair_target_name(anc)
            if repaired and repaired != anc:
                print(f"   Repaired anchor_column '{anc}' → '{repaired}' (target tail-match)")
                result['anchor_column'] = repaired
        elif isinstance(anc, list):
            new_anc = []
            changed = False
            for a in anc:
                r = _repair_target_name(a)
                if r and r != a:
                    changed = True
                new_anc.append(r if r else a)
            if changed:
                print(f"   Repaired composite anchor_column {anc} → {new_anc}")
                result['anchor_column'] = new_anc

        # Repair every column_mappings value. Keep keys (source headers)
        # intact — only target-side values can be wrong here.
        repaired_mappings = {}
        for src_key, tgt_val in merged.items():
            r = _repair_target_name(tgt_val)
            if tgt_val and not r:
                # Truly unresolvable: log and drop (set to None) so the
                # validation guard at confirm_apply surfaces it instead of
                # writing into a phantom column.
                print(
                    f"   WARN: column_mappings value '{tgt_val}' for source "
                    f"'{src_key}' is not a valid target header and could not "
                    f"be repaired by tail-match — dropping to null"
                )
                repaired_mappings[src_key] = None
            elif r and r != tgt_val:
                print(f"   Repaired column_mappings['{src_key}'] '{tgt_val}' → '{r}'")
                repaired_mappings[src_key] = r
            else:
                repaired_mappings[src_key] = tgt_val
        result['column_mappings'] = repaired_mappings
        merged = repaired_mappings
        # Surface the pre-resolved section so the writer pins to it via
        # section_override (mirrors the all-matched fast-path above).
        # Without this, the writer's _pick_best_section auto-pick can
        # land on a different section than the one the LLM was scoped
        # to — for source data that fits multiple sections to similar
        # degrees, the auto-pick's tie-breaker can choose the wrong one.
        if matched_section_idx is not None:
            result['target_section_index'] = matched_section_idx
            result['target_section_title'] = (
                matched_section.get('title') if matched_section else None
            )

        # Post-AI structural override: the AI regularly misclassifies
        # cross_tab / horizontal sources as row_per_date or row_per_entity
        # (because a date/id column happens to match). If the target shape
        # clearly indicates a structural layout, override — the pre-AI
        # detector catches most cases, but this is the safety net for when
        # samples weren't available until after AI inspection.
        ai_strategy = result.get('write_strategy')
        if ai_strategy in ('row_per_date', 'row_per_entity', 'composite_key', 'append'):
            structural_override = _detect_cross_tab_layout(target_schema, source_schema) \
                or _detect_horizontal_layout(target_schema, source_schema)
            if structural_override:
                print(
                    f"Overriding AI '{ai_strategy}' → "
                    f"'{structural_override['write_strategy']}' "
                    f"(structural layout detected)"
                )
                merged_with_structural = dict(merged)
                merged_with_structural.update(structural_override.get('column_mappings') or {})
                result['write_strategy'] = structural_override['write_strategy']
                result['anchor_column'] = structural_override.get('anchor_column')
                result['source_anchor'] = structural_override.get('source_anchor')
                result['anchor_type'] = structural_override.get('anchor_type') or 'name'
                result['column_mappings'] = merged_with_structural
                prior_reasoning = result.get('reasoning') or ''
                result['reasoning'] = (
                    prior_reasoning
                    + f" [local-override: {structural_override['write_strategy']} layout detected]"
                ).strip()
                return result

        # Post-AI reconciliation: if AI picked 'append' but local inference on the
        # merged mappings detects a date (or composite) key, override. This
        # prevents silent duplicates when source dates overlap target dates.
        try:
            if result.get('write_strategy') == 'append':
                # First: if the target actually looks like a 2-column
                # label/value layout, override to key_value regardless of
                # what the AI guessed. This catches the common case where
                # a KPI dashboard has the label column header blank (A1
                # empty) and the AI can't match the source label col.
                kv_layout = _detect_key_value_layout(target_schema, source_schema, merged)
                if kv_layout:
                    label_anchor, label_src, _value_src = kv_layout
                    print(f"Overriding AI 'append' → 'key_value' (anchor='{label_anchor}')")
                    result['write_strategy'] = 'key_value'
                    result['anchor_column'] = label_anchor
                    result['source_anchor'] = label_src
                    result['anchor_type'] = 'name'
                    # Ensure the label source is at least present in the
                    # merged mapping (map it to the target label col even if
                    # it was unmatched due to a blank header).
                    if label_src not in merged or not merged.get(label_src):
                        merged[label_src] = label_anchor
                        result['column_mappings'] = merged
                    prior_reasoning = result.get('reasoning') or ''
                    result['reasoning'] = (
                        prior_reasoning + " [local-override: 2-column key/value layout detected]"
                    ).strip()
                else:
                    local_strategy, local_anchor, local_src_anchor, local_anchor_type = \
                        _infer_strategy_local(target_schema, source_schema, merged)
                    if local_strategy in ('row_per_date', 'composite_key') \
                            and local_anchor and local_src_anchor:
                        note = (
                            ' [local-override: composite key detected]'
                            if local_strategy == 'composite_key'
                            else ' [local-override: date key detected]'
                        )
                        print(f"Overriding AI 'append' → '{local_strategy}' (anchor={local_anchor})")
                        result['write_strategy'] = local_strategy
                        result['anchor_column'] = local_anchor
                        result['source_anchor'] = local_src_anchor
                        result['anchor_type'] = local_anchor_type or (
                            'composite' if local_strategy == 'composite_key' else 'date'
                        )
                        prior_reasoning = result.get('reasoning') or ''
                        result['reasoning'] = (prior_reasoning + note).strip()
                    else:
                        # No date anchor found — try entity-overlap detection.
                        # If any mapped non-date column has a meaningful sample
                        # overlap between source and target (e.g. same SKU / Product
                        # / Store names), the AI is almost certainly wrong to pick
                        # 'append' (that would duplicate the row instead of
                        # updating it). Promote to row_per_entity with the best
                        # overlapping column as the anchor.
                        overlap_anchor = _detect_entity_overlap_anchor(
                            target_schema, source_schema, merged
                        )
                        if overlap_anchor:
                            tgt_col, src_col, overlap_ratio = overlap_anchor
                            print(
                                f"Overriding AI 'append' → 'row_per_entity' "
                                f"(anchor='{tgt_col}', overlap={overlap_ratio:.0%})"
                            )
                            result['write_strategy'] = 'row_per_entity'
                            result['anchor_column'] = tgt_col
                            result['source_anchor'] = src_col
                            result['anchor_type'] = 'id'
                            prior_reasoning = result.get('reasoning') or ''
                            result['reasoning'] = (
                                prior_reasoning
                                + f" [local-override: {overlap_ratio:.0%} entity overlap on '{tgt_col}']"
                            ).strip()
            elif result.get('write_strategy') in ('row_per_date', 'row_per_entity'):
                # AI picked a single-column anchor. Two reconciliation passes:
                #   (a) NEW (TC-A02 fix): if AI picked row_per_date but the
                #       merged mappings contain an entity-id column (SKU/ID/
                #       Code/Key/Name) AND the AI's date pick is just an
                #       audit timestamp ("Last Updated") OR an entity column
                #       exists at all on a mostly-non-time-series target,
                #       prefer row_per_entity. Catalog tables that happen to
                #       carry a Date column should never key by Date.
                #   (b) Existing: if the chosen anchor has duplicates in the
                #       target, promote to composite_key.
                ai_anchor = result.get('anchor_column')
                ai_src_anchor = result.get('source_anchor')
                # (a) date -> entity demotion. If the AI keyed on a date but
                # the target also exposes an entity-id column, that entity is
                # almost always the real record key (catalog tables, KPI by
                # SKU, attendance by employee, etc.). row_per_date on such a
                # fixture silently overwrites every row that shares the same
                # date. The subsequent composite-key check (b) will still
                # bump us back to (entity, date) when warranted.
                if (
                    result.get('write_strategy') == 'row_per_date'
                    and isinstance(ai_anchor, str)
                ):
                    entity_pick = None  # (tgt_col, src_col)
                    for src_h, tgt_h in merged.items():
                        if not tgt_h or not src_h:
                            continue
                        n = str(tgt_h).strip().lower()
                        # Skip the AI's own anchor and any other date column.
                        if tgt_h == ai_anchor or 'date' in n:
                            continue
                        if any(kw in n for kw in _ENTITY_ID_HINTS):
                            entity_pick = (tgt_h, src_h)
                            break
                    if entity_pick:
                        new_tgt, new_src = entity_pick
                        old_anchor = ai_anchor
                        print(
                            f"Overriding AI 'row_per_date' (anchor='{old_anchor}') "
                            f"-> 'row_per_entity' (anchor='{new_tgt}') because "
                            f"mappings carry an entity-id column"
                        )
                        result['write_strategy'] = 'row_per_entity'
                        result['anchor_column'] = new_tgt
                        result['source_anchor'] = new_src
                        result['anchor_type'] = 'id'
                        ai_anchor = new_tgt
                        ai_src_anchor = new_src
                        prior_reasoning = result.get('reasoning') or ''
                        result['reasoning'] = (
                            prior_reasoning
                            + f" [local-override: prefer entity anchor '{new_tgt}' over date '{old_anchor}']"
                        ).strip()
                # (b) single-column -> composite if duplicates
                if isinstance(ai_anchor, str) and isinstance(ai_src_anchor, str):
                    composite = _detect_composite_anchor(target_schema, merged, ai_anchor)
                    if composite:
                        disambig_tgt, disambig_src = composite
                        print(
                            f"Overriding AI '{result.get('write_strategy')}' → 'composite_key' "
                            f"(anchor=[{ai_anchor}, {disambig_tgt}])"
                        )
                        result['write_strategy'] = 'composite_key'
                        result['anchor_column'] = [ai_anchor, disambig_tgt]
                        result['source_anchor'] = [ai_src_anchor, disambig_src]
                        result['anchor_type'] = 'composite'
                        prior_reasoning = result.get('reasoning') or ''
                        result['reasoning'] = (
                            prior_reasoning
                            + f" [local-override: composite key detected on '{ai_anchor}' + '{disambig_tgt}']"
                        ).strip()
        except Exception as _override_err:
            print(f" Post-AI reconciliation skipped: {_override_err}")

        # Final fuzzy-fill pass. Two failure modes this catches:
        #   1. TC-L06 — a polluted target tab with 100+ headers makes the AI
        #      return null for some non-exact source columns; if those cols
        #      have a clean fuzzy-match candidate (case/punct stripped) we
        #      should prefer that over silently dropping the column.
        #   2. TC-L02 — flat source headers like "Inbound Metrics > Trucks >
        #      Pallets" need to match nested target ``composite_to_col_index``
        #      keys which use a different separator. Leaf-segment match
        #      catches those.
        try:
            new_mappings = _fuzzy_fill_unmatched_mappings(
                target_schema, source_schema, result.get('column_mappings') or {}
            )
            if new_mappings:
                fill_summary = ', '.join(f"{s}->{t}" for s, t in new_mappings.items())
                print(f"Fuzzy fill-in added {len(new_mappings)} mapping(s): {fill_summary}")
                result['column_mappings'] = {
                    **(result.get('column_mappings') or {}),
                    **new_mappings,
                }
                prior_reasoning = result.get('reasoning') or ''
                result['reasoning'] = (
                    prior_reasoning + f" [fuzzy-fill: {len(new_mappings)} extra column(s) matched]"
                ).strip()
        except Exception as _fill_err:
            print(f" Fuzzy fill-in skipped: {_fill_err}")

        print(f"   Strategy: {result.get('write_strategy')}")
        print(f"   Anchor: {result.get('anchor_column')}")
        print(f"   Total mappings: {len(exact_mappings)} exact + {len(ai_mappings)} AI")
        print(f"   Reasoning: {result.get('reasoning')}")
        return result

    except Exception as e:
        print(f"find_identifier error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


# ---------------------------------------------------------------------------
# Fuzzy fill-in helpers
# ---------------------------------------------------------------------------
# After the AI mapping pass and any local strategy overrides, some source
# columns may still be unmapped because the target headers carry punctuation,
# casing, or hierarchy noise that the AI couldn't traverse. This helper does
# a final pass with cheap, deterministic match strategies — no AI calls, no
# remote dependencies, fully unit-testable.

import re as _re_fill


def _strip_for_compare(s: str) -> str:
    """Lowercase, strip non-alphanumerics, collapse whitespace. Used as the
    canonical form for fuzzy header comparison."""
    if not s:
        return ''
    cleaned = _re_fill.sub(r'[^a-z0-9]+', ' ', str(s).lower()).strip()
    return _re_fill.sub(r'\s+', ' ', cleaned)


def _split_composite_path(s: str):
    """Split a composite header path on the common nested-header separators
    ('>', '|', ' / ', ' :: '). Returns the leaf segment plus the full list
    of segments (both stripped). Returns (s, [s]) when no separator is
    present."""
    if not s:
        return '', []
    raw = str(s)
    for sep in (' > ', '>', ' | ', '|', ' / ', '/', ' :: ', '::'):
        if sep in raw:
            parts = [p.strip() for p in raw.split(sep) if p.strip()]
            if len(parts) >= 2:
                return parts[-1], parts
    return raw.strip(), [raw.strip()]


def _fuzzy_fill_unmatched_mappings(target_schema, source_schema, current_mappings):
    """Return a dict of NEW source->target mappings for source columns not
    currently in ``current_mappings`` (or mapped to None/empty). Only emits
    a mapping when there's exactly one unambiguous fuzzy candidate; ties
    are skipped because guessing wrong overwrites the wrong target column.

    Match strategies, in priority order:
        1. case/punct-stripped equality of the full source header against
           every target header.
        2. Same comparison against every key in target ``composite_to_
           col_index`` (handles nested target headers).
        3. Leaf-segment equality: split source on ' > ' / '|' / '/', compare
           the trailing segment against target headers and composite keys.
        4. Strict containment in either direction (source ⊆ target OR
           target ⊆ source) on the stripped form, gated by length >= 4 to
           avoid pathological short-token collisions.

    The mapper deliberately stops before any "approximate" match (Levenshtein
    etc.) — the AI is supposed to handle that, and silent fuzzy guesses are
    riskier than leaving a column unmapped (the UI will surface it).
    """
    if not current_mappings:
        current_mappings = {}

    target_headers = list(target_schema.get('headers') or [])
    composite_keys = list((target_schema.get('composite_to_col_index') or {}).keys())
    formula_cols = set(target_schema.get('formula_cols') or [])

    # Source columns still needing a target. Mappings with falsy values count
    # as unmapped so AI nulls get a second chance here.
    source_headers = list(source_schema.get('headers') or [])
    used_targets = {v for v in current_mappings.values() if v}
    unmatched_sources = [
        s for s in source_headers
        if s not in current_mappings or not current_mappings.get(s)
    ]
    if not unmatched_sources:
        return {}

    # Build a normalized lookup from canonical form -> target header.
    # Multiple targets normalizing to the same key get collapsed into a list
    # so we can detect ambiguity and skip those.
    norm_to_targets = {}
    def _record(canon: str, target: str):
        if not canon or not target:
            return
        if target in formula_cols:
            return
        norm_to_targets.setdefault(canon, []).append(target)

    for h in target_headers:
        _record(_strip_for_compare(h), h)
    for ck in composite_keys:
        _record(_strip_for_compare(ck), ck)
        leaf, _ = _split_composite_path(ck)
        if leaf and leaf != ck:
            _record(_strip_for_compare(leaf), ck)

    new_mappings = {}
    for src in unmatched_sources:
        canon_src = _strip_for_compare(src)
        leaf_src, _ = _split_composite_path(src)
        canon_leaf = _strip_for_compare(leaf_src) if leaf_src else canon_src

        candidate = None

        # 1+2: exact canonical match on full or composite-key form.
        for key in (canon_src, canon_leaf):
            if not key:
                continue
            hits = norm_to_targets.get(key, [])
            # Filter out targets already claimed by another source.
            hits = [t for t in hits if t not in used_targets]
            if len(hits) == 1:
                candidate = hits[0]
                break
            if len(hits) > 1:
                # Ambiguous — don't guess.
                candidate = None
                break

        # 3: stripped substring containment (gated to long-ish tokens).
        if candidate is None and len(canon_src) >= 4:
            sub_hits = []
            for key, tgts in norm_to_targets.items():
                if not key or len(key) < 4:
                    continue
                if canon_src == key:
                    continue
                if canon_src in key or key in canon_src:
                    for t in tgts:
                        if t not in used_targets and t not in sub_hits:
                            sub_hits.append(t)
            if len(sub_hits) == 1:
                candidate = sub_hits[0]

        if candidate:
            new_mappings[src] = candidate
            used_targets.add(candidate)

    return new_mappings


def _detect_entity_overlap_anchor(target_schema, source_schema, mappings,
                                min_overlap=0.5, min_samples=2):
    """
    Return (target_col, source_col, overlap_ratio) for the mapped non-date
    column whose source/target sample values overlap the most (>= min_overlap),
    or None if no mapping qualifies.

    This powers the reconciliation that promotes an AI-picked 'append' to
    'row_per_entity' when the source is clearly an UPDATE to existing rows
    (same SKUs, same products, same store names, etc.), not a fresh append.
    """
    # structure_*_data exposes sample values under 'col_samples' (see
    # structure_source_data / structure_target_data above).
    target_samples = target_schema.get('col_samples', {}) or target_schema.get('samples', {}) or {}
    source_samples = source_schema.get('col_samples', {}) or source_schema.get('samples', {}) or {}
    target_types = target_schema.get('col_types', {}) or {}
    source_types = source_schema.get('col_types', {}) or {}

    best = None  # (ratio, tgt_col, src_col)

    for src_col, tgt_col in mappings.items():
        if not tgt_col:
            continue

        name_lower = str(tgt_col).lower()
        # Dates handled by the other override. Numerics are almost never
        # identifiers; overlap there is coincidence (prices, quantities).
        if 'date' in name_lower:
            continue
        if 'date' in str(source_types.get(src_col, '')).lower():
            continue
        if 'date' in str(target_types.get(tgt_col, '')).lower():
            continue
        if str(target_types.get(tgt_col, '')).lower() == 'numeric':
            continue
        if str(source_types.get(src_col, '')).lower() == 'numeric':
            continue

        src_vals = source_samples.get(src_col) or []
        tgt_vals = target_samples.get(tgt_col) or []
        src_norm = {str(v).strip().lower() for v in src_vals if v not in (None, '')}
        tgt_norm = {str(v).strip().lower() for v in tgt_vals if v not in (None, '')}

        if len(src_norm) < min_samples or len(tgt_norm) < min_samples:
            continue

        intersection = src_norm & tgt_norm
        if not intersection:
            continue

        ratio = len(intersection) / max(len(src_norm), 1)

        # Prefer id/code/name-ish columns when tie-breaking.
        keyword_boost = 1.0 if any(kw in name_lower for kw in (
            'id', 'sku', 'code', 'key', 'name', 'product', 'store', 'customer',
            'entity', 'item'
        )) else 0.0
        score = ratio + keyword_boost * 0.01

        if ratio >= min_overlap and (best is None or score > best[0]):
            best = (score, tgt_col, src_col, ratio)

    if best is None:
        return None
    _, tgt_col, src_col, ratio = best
    return tgt_col, src_col, ratio


def _detect_composite_anchor(target_schema, mappings, primary_tgt_col):
    """
    Decide whether the target needs a composite key anchored on
    ``primary_tgt_col`` plus one more mapped column.

    Scans every data row of ``target_schema['raw_rows']`` (not just the
    5-value sample in ``col_samples``) so duplicate detection is accurate
    on full sheets. Returns ``(disambig_tgt_col, disambig_src_col)`` when a
    suitable second key column exists, or ``None`` when the primary anchor
    is already unique / no good disambiguator is available.

    A disambiguator is considered suitable when:
    - it's a different mapped target column that isn't a formula, isn't
        another date column, and isn't purely numeric (measurements like
        Sales/Qty are coincidentally unique but aren't identifiers);
    - the resulting ``(primary, disambiguator)`` pairs are substantially
        more unique than primary alone (ratio >= 0.9 of unique / total).

    Columns whose names hint at an identifier role
    (store / shift / product / region / branch / sku / id / name / code ...)
    get a small score boost so we prefer ``Date + Store`` over
    ``Date + Notes``.
    """
    raw_rows = target_schema.get('raw_rows') or []
    header_row_count = target_schema.get('header_row_count', 1) or 1
    header_index = target_schema.get('header_index') or {}
    formula_cols = set(target_schema.get('formula_cols') or [])
    target_types = target_schema.get('col_types', {}) or {}

    if primary_tgt_col not in header_index:
        return None

    data_rows = raw_rows[header_row_count:] if len(raw_rows) > header_row_count else []
    if len(data_rows) < 2:
        return None

    primary_idx = header_index[primary_tgt_col]
    primary_vals = []
    for r in data_rows:
        v = r[primary_idx] if primary_idx < len(r) else None
        if v is not None and str(v).strip():
            primary_vals.append(str(v).strip().lower())

    if len(primary_vals) < 2:
        return None
    if len(set(primary_vals)) == len(primary_vals):
        return None

    preferred_kw = (
        'store', 'shift', 'product', 'region', 'branch', 'location',
        'outlet', 'warehouse', 'sku', 'item', 'id', 'code', 'name',
        'category', 'dept', 'department', 'team', 'segment', 'channel',
    )

    best = None  # (score, tgt, src)
    for src_col, tgt_col in mappings.items():
        if not tgt_col or tgt_col == primary_tgt_col:
            continue
        if tgt_col in formula_cols or tgt_col not in header_index:
            continue

        tgt_type = str(target_types.get(tgt_col, '')).lower()
        name_lower = str(tgt_col).lower()
        if tgt_type == 'numeric':
            continue
        if 'date' in tgt_type or 'date' in name_lower:
            continue

        idx = header_index[tgt_col]
        pairs = []
        for r in data_rows:
            pv = r[primary_idx] if primary_idx < len(r) else None
            dv = r[idx] if idx < len(r) else None
            pv_s = str(pv).strip().lower() if pv is not None and str(pv).strip() else ''
            dv_s = str(dv).strip().lower() if dv is not None and str(dv).strip() else ''
            if pv_s:
                pairs.append((pv_s, dv_s))

        if len(pairs) < 2:
            continue

        uniq_ratio = len(set(pairs)) / len(pairs)
        keyword_boost = 0.1 if any(kw in name_lower for kw in preferred_kw) else 0.0
        score = uniq_ratio + keyword_boost
        if best is None or score > best[0]:
            best = (score, tgt_col, src_col)

    if best is None:
        return None
    score, tgt_col, src_col = best
    if score < 0.9:
        return None
    return (tgt_col, src_col)


def _detect_key_value_layout(target_schema, source_schema, mappings):
    """Return (label_tgt_col, label_src_col) if the target/source pair looks
    like a 2-column label/value (metric/value) layout, else None.

    Heuristic: the target has exactly 2 columns (ignoring blanks), the first
    column is string-typed and the second is numeric / mixed (the value), and
    at least one mapping lands on each column. This covers typical KPI
    dashboards like (Metric, Value), (Label, Amount), (KPI, Total), etc.
    """
    tgt_headers = [h for h in target_schema.get('headers', []) if h is not None]
    if len([h for h in tgt_headers if str(h).strip() != '']) > 2:
        return None
    # Need exactly two target slots (blank A1 counts as the label slot).
    if len(tgt_headers) != 2:
        return None

    target_types = target_schema.get('col_types', {}) or {}
    label_col, value_col = tgt_headers[0], tgt_headers[1]
    label_type = str(target_types.get(label_col, '')).lower()
    value_type = str(target_types.get(value_col, '')).lower()

    # Infer from raw rows if type dict is empty (e.g. blank header column).
    if not label_type or not value_type:
        raw = target_schema.get('raw_rows', [])
        if len(raw) >= 2:
            samples_a, samples_b = [], []
            for r in raw[1:]:
                if len(r) > 0 and r[0] not in (None, ''):
                    samples_a.append(str(r[0]).strip())
                if len(r) > 1 and r[1] not in (None, ''):
                    samples_b.append(str(r[1]).strip())
            label_type = label_type or _infer_col_type(samples_a)
            value_type = value_type or _infer_col_type(samples_b)

    # Label column must be string-like; value column must NOT be string-only.
    # (A numeric or mixed value column is what makes it key_value vs a generic
    # 2-col entity table.)
    if 'string' not in label_type and 'text' not in label_type:
        return None
    if value_type and 'string' in value_type and 'number' not in value_type:
        return None

    # Find the matching source columns for label + value.
    label_src = None
    value_src = None
    for src, tgt in mappings.items():
        if tgt == label_col:
            label_src = src
        elif tgt == value_col:
            value_src = src
    # If the label header is blank (A1 empty) the source label col probably
    # went unmapped — pick any remaining source column as the label.
    if not label_src:
        mapped_srcs = {s for s, t in mappings.items() if t}
        for src in source_schema.get('headers', []):
            if src not in mapped_srcs:
                label_src = src
                break
    if not value_src:
        return None
    if not label_src:
        return None

    # When the target's label header is blank (A1 empty), we can't use ''
    # as the anchor (the UI would show an empty badge AND the write-side
    # lookup would then match '' against 'Value' partially). Use the source
    # label column name instead — _write_key_value looks up the header name
    # in the target; a miss falls back to label_idx=0 / value_idx=1, which
    # is exactly what we want for this layout.
    anchor = label_col if str(label_col).strip() else label_src
    return (anchor, label_src, value_src)


def _detect_cross_tab_layout(target_schema, source_schema):
    """Detect cross_tab when a source column's VALUES substantially match
    target column HEADERS (pivot signal). Returns a full identification dict
    or None.

    Example trigger:
    source headers = [Date, Category, Revenue]
    source Category values = {Electronics, Apparel}
    target headers = [Date, Electronics, Apparel, Grocery]
    → cross_tab (anchor=Date, pivot_src=Category, value_src=Revenue)
    """
    target_headers = [h for h in (target_schema.get('headers') or []) if h]
    source_headers = source_schema.get('headers') or []
    if len(target_headers) < 3 or len(source_headers) < 3:
        return None

    tgt_first = target_headers[0]
    tgt_data_headers_lower = {
        h.strip().lower() for h in target_headers[1:] if h and str(h).strip()
    }
    if len(tgt_data_headers_lower) < 2:
        return None

    col_samples = source_schema.get('col_samples', {}) or {}

    best = None
    for src_col in source_headers:
        samples = col_samples.get(src_col, []) or []
        if not samples:
            continue
        sample_set = {str(s).strip().lower() for s in samples if str(s).strip()}
        if len(sample_set) < 2:
            continue
        overlap = sample_set & tgt_data_headers_lower
        if len(overlap) < 2:
            continue
        overlap_ratio = len(overlap) / len(sample_set)
        if overlap_ratio < 0.5:
            continue
        score = overlap_ratio + 0.05 * len(overlap)
        if best is None or score > best[0]:
            best = (score, src_col, overlap_ratio, overlap)

    if not best:
        return None

    _score, pivot_src, overlap_ratio, overlap = best

    # Row anchor resolution: match target[0] to a source column by name,
    # else fall back to a date-typed source column, else first remaining.
    remaining_src = [c for c in source_headers if c != pivot_src]
    tgt_first_lower = str(tgt_first).strip().lower()
    row_src = None
    for c in remaining_src:
        if c.strip().lower() == tgt_first_lower:
            row_src = c
            break
    if not row_src:
        src_types = source_schema.get('col_types', {}) or {}
        for c in remaining_src:
            if 'date' in src_types.get(c, '').lower() or 'date' in c.lower():
                row_src = c
                break
    if not row_src and remaining_src:
        row_src = remaining_src[0]
    if not row_src:
        return None

    # Value column: remaining non-pivot non-anchor source column
    value_src = None
    for c in remaining_src:
        if c != row_src:
            value_src = c
            break

    # The orchestrator will pre-pivot source rows into wide form using
    # pivot_source_col + value_source_col, then rebuild column_mappings as
    # identity (anchor + each pivot value that matches a target column).
    # At detection time we return only the anchor mapping; the real column
    # mappings are produced post-pivot.
    mappings = {row_src: tgt_first}

    anchor_type = 'date' if 'date' in str(row_src).lower() else 'name'
    src_types = source_schema.get('col_types', {}) or {}
    if 'date' in src_types.get(row_src, '').lower():
        anchor_type = 'date'

    return {
        'write_strategy': 'cross_tab',
        'anchor_column': tgt_first,
        'source_anchor': row_src,
        'anchor_type': anchor_type,
        'column_mappings': mappings,
        'pivot_source_col': pivot_src,
        'value_source_col': value_src,
        'reasoning': (
            f"Source '{pivot_src}' values overlap {overlap_ratio:.0%} with "
            f"target column headers ({len(overlap)}/{len(tgt_data_headers_lower)} match). "
            f"Pivoting '{value_src}' values into cross_tab matrix anchored on '{tgt_first}'."
        ),
    }


def _detect_horizontal_layout(target_schema, source_schema):
    """Detect horizontal when target column headers are time periods
    (Jan..Dec, Q1..Q4, weekdays, 4-digit years, Wn) AND the source has at
    least two matching period columns. Returns a full identification dict
    or None.
    """
    import re as _re

    target_headers = [h for h in (target_schema.get('headers') or []) if h]
    source_headers = source_schema.get('headers') or []
    if len(target_headers) < 3 or len(source_headers) < 2:
        return None

    MONTHS = {
        'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug',
        'sep', 'oct', 'nov', 'dec',
        'january', 'february', 'march', 'april', 'june', 'july',
        'august', 'september', 'october', 'november', 'december',
    }
    QUARTERS = {'q1', 'q2', 'q3', 'q4'}
    DAYS = {
        'mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun',
        'monday', 'tuesday', 'wednesday', 'thursday',
        'friday', 'saturday', 'sunday',
    }

    def is_period(h):
        hl = str(h).strip().lower()
        if not hl:
            return False
        if hl in MONTHS or hl in QUARTERS or hl in DAYS:
            return True
        if _re.match(r'^\d{4}$', hl):  # year
            return True
        if _re.match(r'^w\d{1,2}$', hl):  # week number
            return True
        return False

    period_tgt = [h for h in target_headers if is_period(h)]
    if len(period_tgt) < 3:
        return None

    period_tgt_lower = {h.strip().lower() for h in period_tgt}
    period_src = [h for h in source_headers if h.strip().lower() in period_tgt_lower]
    if len(period_src) < 2:
        return None

    # Row anchor: first target header that is NOT a period.
    row_anchor = None
    for h in target_headers:
        if not is_period(h):
            row_anchor = h
            break

    row_src = None
    if row_anchor:
        row_anchor_lower = str(row_anchor).strip().lower()
        for s in source_headers:
            if s.strip().lower() == row_anchor_lower:
                row_src = s
                break
        if not row_src:
            for s in source_headers:
                if not is_period(s):
                    row_src = s
                    break

    mappings = {}
    if row_src and row_anchor:
        mappings[row_src] = row_anchor
    for p in period_src:
        pl = p.strip().lower()
        for t in period_tgt:
            if t.strip().lower() == pl:
                mappings[p] = t
                break

    return {
        'write_strategy': 'horizontal',
        'anchor_column': row_anchor,
        'source_anchor': row_src,
        'anchor_type': 'name',
        'column_mappings': mappings,
        'reasoning': (
            f"Target has {len(period_tgt)} time-period columns "
            f"({', '.join(period_tgt[:4])}{'...' if len(period_tgt) > 4 else ''}); "
            f"source provides {len(period_src)} matching periods. "
            f"Using horizontal layout with row anchor '{row_anchor}'."
        ),
    }


def _detect_structural_layout(target_schema, source_schema, exact_mappings):
    """Run all structural-layout detectors in priority order and return the
    first match. Checks BEFORE the AI call to avoid the "row_per_date
    fallback" misclassification pattern.

    Order (most-to-least specific):
    1. cross_tab  — source values match target headers
    2. horizontal — target headers are time periods
    3. key_value  — 2-column label/value target
    """
    hit = _detect_cross_tab_layout(target_schema, source_schema)
    if hit:
        return hit

    hit = _detect_horizontal_layout(target_schema, source_schema)
    if hit:
        return hit

    kv = _detect_key_value_layout(target_schema, source_schema, exact_mappings)
    if kv:
        label_anchor, label_src, value_src = kv
        mappings = dict(exact_mappings)
        if label_src not in mappings or not mappings.get(label_src):
            mappings[label_src] = label_anchor
        return {
            'write_strategy': 'key_value',
            'anchor_column': label_anchor,
            'source_anchor': label_src,
            'anchor_type': 'name',
            'column_mappings': mappings,
            'reasoning': '2-column label/value layout detected.',
        }

    return None


# Substrings that mark a date column as METADATA (audit timestamp) rather
# than a primary record key. A column literally named "Last Updated",
# "Modified", "Created At", etc. should never be picked as the row anchor —
# treat it like any other data column. Matched against the lowercased target
# header.
_SECONDARY_DATE_HINTS = (
    'last updated', 'last modified', 'last edited',
    'modified', 'updated', 'created', 'edited',
    'modified_at', 'updated_at', 'created_at',
    'modified at', 'updated at', 'created at',
    'timestamp', 'modified date', 'updated date', 'created date',
)


def _is_primary_date_column(name: str) -> bool:
    """Return True when ``name`` looks like a real date anchor (e.g. "Date",
    "Order Date", "Sales Date") rather than an audit timestamp. The check is
    deliberately conservative: if 'date' is in the name and none of the
    secondary hints match, it's a primary candidate."""
    if not name:
        return False
    n = name.strip().lower()
    if 'date' not in n:
        return False
    return not any(hint in n for hint in _SECONDARY_DATE_HINTS)


# Header keywords that mark a column as an entity identifier. Order matters
# only for documentation; the check is a substring scan so 'sku' matches
# 'SKU Code', 'product_id' matches 'Product ID', etc.
_ENTITY_ID_HINTS = ('id', 'sku', 'code', 'key', 'name')


def _infer_strategy_local(target_schema, source_schema, mappings):
    """Infer write strategy without AI when all columns matched exactly.

    Anchor preference order (changed 2026-05 to fix TC-A02-style fixtures
    where an entity table also carries an updated_at-style date column):
        1. multi_section target            (writes per detected section)
        2. key_value 2-col label/value     (cheaper than full row analysis)
        3. entity ID columns               (sku/id/code/key/name)
        4. primary date columns            (excluding "Last Updated" etc.)
        5. append                          (no anchor inferable)

    When BOTH an entity ID and a primary date are mapped, ``_detect_composite_
    anchor`` is consulted for either path so a (SKU, Date) ledger correctly
    promotes to ``composite_key`` regardless of which loop matched first.
    """
    # If multi-section target, use multi_section strategy
    sections = target_schema.get('sections', [])
    if sections:
        for src, tgt in mappings.items():
            src_type = source_schema.get('col_types', {}).get(src, '')
            if _is_primary_date_column(tgt) or 'date' in src_type:
                return 'multi_section', tgt, src, 'date'
        first_col = next(iter(mappings.values()), None)
        first_src = next(iter(mappings.keys()), None)
        return 'multi_section', first_col, first_src, 'name'

    # 2-column label/value layout (key_value). Checked before the
    # date / id heuristics so a target like (Metric, Value) with numeric
    # values isn't misclassified as row_per_entity just because "Metric"
    # matched the 'name' keyword.
    kv = _detect_key_value_layout(target_schema, source_schema, mappings)
    if kv:
        label_anchor, label_src, _value_src = kv
        return 'key_value', label_anchor, label_src, 'name'

    target_types = target_schema.get('col_types', {})
    source_types = source_schema.get('col_types', {})

    # Entity ID anchor scan FIRST (TC-A02 fix). When a fixture has both SKU
    # and Date, SKU is the real record identity; Date in that context is
    # usually a mutation timestamp, not the partition key. row_per_date on
    # such a fixture silently overwrites every row that shares a date.
    for src, tgt in mappings.items():
        name_lower = tgt.lower() if tgt else ''
        if any(kw in name_lower for kw in _ENTITY_ID_HINTS):
            composite = _detect_composite_anchor(target_schema, mappings, tgt)
            if composite:
                disambig_tgt, disambig_src = composite
                return (
                    'composite_key',
                    [tgt, disambig_tgt],
                    [src, disambig_src],
                    'composite',
                )
            return 'row_per_entity', tgt, src, 'id'

    # Primary date anchor scan (only after entity scan came up empty).
    # _is_primary_date_column filters out "Last Updated" / "Modified Date"
    # so an audit column doesn't masquerade as the row anchor.
    for src, tgt in mappings.items():
        src_type = source_types.get(src, '')
        tgt_type = target_types.get(tgt, '')
        if _is_primary_date_column(tgt) or 'date' in src_type or 'date' in tgt_type:
            composite = _detect_composite_anchor(target_schema, mappings, tgt)
            if composite:
                disambig_tgt, disambig_src = composite
                return (
                    'composite_key',
                    [tgt, disambig_tgt],
                    [src, disambig_src],
                    'composite',
                )
            return 'row_per_date', tgt, src, 'date'

    return 'append', None, None, None
# ============================================================
# TOOL REGISTRY
# ============================================================
TOOL_REGISTRY = {
    "parse_file": {
        "func": parse_file,
        "description": "Parse CSV/Excel/JSON files into structured data",
    },
    "detect_source_sections": {
        "func": detect_source_sections,
        "description": "Detect multiple stacked sections (title+header+data) inside a source xlsx file",
    },
    "detect_source_sheets": {
        "func": detect_source_sheets,
        "description": "Enumerate sheets in a source xlsx and score each by target-header overlap",
    },
    "extract_dates_from_all_rows": {
        "func": extract_dates_from_all_rows,
        "description": "Extract dates from all rows for date-based matching",
    },
    "smart_column_mapping": {
        "func": smart_column_mapping,
        "description": "Intelligently map source to target columns with AI",
    },
    "transform_data": {
        "func": transform_data,
        "description": "Apply mappings and transform data structure",
    },
    "merge_dates_and_transformed_data": {
        "func": merge_dates_and_transformed_data,
        "description": "Merge dates with transformed data for OPR workflow",
    },
    "structure_target_data": {
        "func": structure_target_data,
        "description": "Convert raw sheet values into a structured schema object",
    },
    "structure_source_data": {
        "func": structure_source_data,
        "description": "Convert parsed file output into a structured schema object",
    },
    "find_identifier": {
        "func": find_identifier,
        "description": "AI call to identify anchor column and map source to target columns",
    },
}


# ============================================================
# API ENDPOINTS
# ============================================================

@app.post("/execute_task", response_model=ToolResponse)
async def execute_tool(request: ToolRequest):
    """Execute a mapping tool"""
    try:
        print(f"\nMapping Agent - Tool: {request.tool}")
        print(f"   Inputs keys: {list(request.inputs.keys())}")
        
        has_credentials = request.credentials is not None
        print(f"   Credentials provided: {has_credentials}")
        
        tool_info = TOOL_REGISTRY.get(request.tool)
        if not tool_info:
            available_tools = list(TOOL_REGISTRY.keys())
            return ToolResponse(
                success=False,
                error=f"Unknown tool: {request.tool}. Available: {available_tools}",
            )
        
        tool_inputs = request.inputs.copy()
        if request.credentials:
            tool_inputs['credentials'] = request.credentials
        
        result = tool_info["func"](**tool_inputs)
        
        print(f"   {'OK' if result.get('success') else 'FAIL'} Result: {result.get('success', False)}")
        
        return ToolResponse(
            success=result.get("success", False),
            result=result if result.get("success") else None,
            error=result.get("error") if not result.get("success") else None,
        )
        
    except Exception as e:
        print(f"   Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return ToolResponse(
            success=False,
            error=f"Tool execution failed: {str(e)}",
        )


@app.get("/tools")
async def list_tools():
    """List all available tools"""
    return {
        "tools": [
            {"name": name, "description": info["description"]}
            for name, info in TOOL_REGISTRY.items()
        ],
        "count": len(TOOL_REGISTRY),
    }


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "mapping-agent",
        "version": "1.0.0",
        "smart_mapping_available": SMART_MAPPING_AVAILABLE,
    }


@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "service": "Mapping Agent API",
        "version": "1.0.0",
        "description": "Data intelligence and transformation microservice",
        "features": [
            "File parsing (CSV, Excel, JSON)",
            "AI-powered column mapping",
            "Data validation",
            "Data transformation",
            "Date extraction for OPR workflow",
            "Template management",
        ],
        "endpoints": {
            "execute": "/execute (POST) - Execute a mapping tool",
            "tools": "/tools (GET) - List available tools",
            "health": "/health (GET) - Health check",
            "docs": "/docs (GET) - Swagger documentation",
        },
    }